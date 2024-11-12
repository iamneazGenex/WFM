from django.shortcuts import render
from colorama import Fore
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import load_workbook
from accounts.models import *
from roster.models import *
from .forms import *
from django.contrib import messages
from django_datatables_view.base_datatable_view import BaseDatatableView
import logging
from django.contrib import messages
from django.db.models import Q
from rms.global_utilities import *
from .utils import *
from django.utils.html import format_html
from rms.page_info_collection import PageInfoCollection
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from rms.constants import GroupEnum
from rms.decorators import check_user_able_to_see_page
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime
from django.db.models import Sum, F
from django.http import JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
import json

logger = logging.getLogger(__name__)


# Create your views here.
################################################################
#   Avaya CDR
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def viewAvayaCDR(request):

    templateName = "avayaCDR/view.html"
    breadCrumbList = [
        PageInfoCollection.AVAYACDR_VIEW,
    ]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AVAYACDR_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Shift Legend",
        },
        "ajaxUrl": PageInfoCollection.SHIFTLEGEND_JSON.urlName,
        "createUrl": PageInfoCollection.SHIFTLEGEND_CREATE.urlName,
        "bulkUrl": PageInfoCollection.AVAYACDR_BULK.urlName,
        "avayaCDRActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def createAvayaCDR(request):
    templateName = "avayaCDR/create.html"
    breadCrumbList = [
        PageInfoCollection.AVAYACDR_VIEW,
        PageInfoCollection.AVAYACDR_CREATE,
    ]
    if request.method == "POST":
        form = AvayaCDRForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = avayaCDRCreation(data, request)
            if result == True:
                messages.success(request, "Shift Legend created successfully.")
                return redirect(PageInfoCollection.SHIFTLEGEND_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Shift Legend")
        else:
            messages.error(request, "Could not create Shift Legend")
    else:
        form = AvayaCDRForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AVAYACDR_CREATE.pageName,
        "formUrl": PageInfoCollection.AVAYACDR_CREATE.urlName,
        "avayaCDRActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def createBulkAvayaCDR(request):
    templateName = "avayaCDR/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.AVAYACDR_VIEW,
        PageInfoCollection.AVAYACDR_BULK,
    ]
    if request.method == "POST":
        sheetName = request.POST.get("sheetName").strip()
        if not sheetName:
            messages.error(request, "Sheet name is empty")
        else:
            try:
                excel_file = request.FILES["excel_file"]
                wb = load_workbook(excel_file, data_only=True)
                ws = wb[sheetName]
                count = (
                    len(
                        [
                            row
                            for row in ws
                            if not all([cell.value == None for cell in row])
                        ]
                    )
                    - 1
                )
                successCount = 0

                for index, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=2
                ):
                    logger.info(f"Row {index}: {row}")
                    if row[0] is not None:
                        try:
                            skill = Skill.objects.get(name=row[1].lower())

                            data = {
                                "date": row[0],
                                "skill": skill,
                                "time_interval": row[2],
                                "hour": 0,  # Set the initial value for hour, you can extract it later
                                "offer": row[3],
                                "answer": row[4],
                                "ans_th": row[5],
                                "aban_calls": row[6],
                                "abn_th": row[7],
                                "avg_acd_time": row[8],
                                "avg_hold_time": row[9],
                                "avg_acw_time": row[10],
                                "avg_speed_ans": row[11],
                                "staffed_time": row[12],
                                "acd_time": row[13],
                                "acw_time": row[14],
                                "hold_time": row[15],
                                "avail_time": row[16],
                                "aux_time": row[17],
                                "agent_ring_time": row[18],
                                "other_time": row[19],
                            }
                            result = avayaCDRCreation(data, request)
                            if result == True:
                                successCount += 1
                            else:
                                messages.error(request, f"Could not create")
                        except Exception as e:
                            logger.info(f"|Failed|Upload failed.Exception: {e}")
                if count == successCount:
                    messages.success(request, "ALL Avaya CDR Uploaded Successfully")
                elif successCount == 0:
                    messages.error(request, "Failed to Upload any Avaya CDR")
                else:
                    messages.warning(request, "Failed to upload some Avaya CDR")
            except Exception as e:
                logger.info(f"|Failed| Bulk Avaya CDR Upload failed.Exception: {e}")
                messages.error(request, "Bulk Avaya CDR Upload failed")
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AVAYACDR_BULK.pageName,
        "formUrl": PageInfoCollection.AVAYACDR_BULK.urlName,
        "avayaCDRActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


################################################################
#   Agent Hourly Performace
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def viewAgentHourlyPerformance(request):

    templateName = "agentHourlyPerformance/view.html"
    breadCrumbList = [
        PageInfoCollection.AGENTHOURLYPERFORMANCE_VIEW,
    ]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AGENTHOURLYPERFORMANCE_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Shift Legend",
        },
        "ajaxUrl": PageInfoCollection.SHIFTLEGEND_JSON.urlName,
        "createUrl": PageInfoCollection.SHIFTLEGEND_CREATE.urlName,
        "bulkUrl": PageInfoCollection.AGENTHOURLYPERFORMANCE_BULK.urlName,
        "agentHourlyPerformanceActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def createAgentHourlyPerformance(request):
    templateName = "avayaCDR/create.html"
    breadCrumbList = [
        PageInfoCollection.AVAYACDR_VIEW,
        PageInfoCollection.AVAYACDR_CREATE,
    ]
    if request.method == "POST":
        form = AvayaCDRForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = avayaCDRCreation(data, request)
            if result == True:
                messages.success(request, "Shift Legend created successfully.")
                return redirect(PageInfoCollection.SHIFTLEGEND_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Shift Legend")
        else:
            messages.error(request, "Could not create Shift Legend")
    else:
        form = AvayaCDRForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AVAYACDR_CREATE.pageName,
        "formUrl": PageInfoCollection.AVAYACDR_CREATE.urlName,
        "avayaCDRActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def createBulkAgentHourlyPerformance(request):
    templateName = "agentHourlyPerformance/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.AGENTHOURLYPERFORMANCE_VIEW,
        PageInfoCollection.AGENTHOURLYPERFORMANCE_BULK,
    ]
    if request.method == "POST":
        sheetName = request.POST.get("sheetName").strip()
        if not sheetName:
            messages.error(request, "Sheet name is empty")
        else:
            try:
                excel_file = request.FILES["excel_file"]
                wb = load_workbook(excel_file, data_only=True)
                ws = wb[sheetName]
                count = (
                    len(
                        [
                            row
                            for row in ws
                            if not all([cell.value == None for cell in row])
                        ]
                    )
                    - 1
                )
                successCount = 0

                for index, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=2
                ):
                    logger.info(f"Row {index}: {row}")
                    if row[0] is not None:
                        try:
                            try:
                                employee = Employee.objects.get(avaya_id=row[1])
                            except Exception as e:
                                logger.info(f"Exception: {e}")
                                employee = None
                            skill = Skill.objects.get(name=row[2].lower())
                            # django_date_format = row[0].strftime("%m/%d/%Y")
                            # converted_date = datetime.strptime(
                            #     django_date_format, "%d/%m/%Y"
                            # )
                            data = {
                                "date": row[0],
                                "employee": employee,
                                "skill": skill,
                                "time_interval_start": row[3],
                                "time_interval_end": row[5],
                                "hour": 0,
                                "acd_calls": row[6],
                                "avg_acd_time": row[7],
                                "avg_acw_time": row[8],
                                "percent_agent_occup_with_acw": row[9],
                                "percent_agent_occup_without_acw": row[10],
                                "extn_in_calls": row[11],
                                "avg_extn_in_time": row[12],
                                "extn_out_calls": row[13],
                                "avg_extn_out_time": row[14],
                                "acd_time": row[15],
                                "acw_time": row[16],
                                "agent_ring_time": row[17],
                                "other_time": row[18],
                                "aux_time": row[19],
                                "avail_time": row[20],
                                "percent_skills_avail": row[21],
                                "staffed_time": row[22],
                                "trans_out": row[23],
                                "held_calls": row[24],
                                "avg_hold_time": row[25],
                                "aux_work_on_login": row[26],
                                "short_break": row[27],
                                "lunch_break": row[28],
                                "training": row[29],
                                "one_to_one": row[30],
                                "meeting": row[31],
                                "outbound_callback": row[32],
                                "cfs_meeting": row[33],
                                "hour_8": row[34],
                                "hour_9": row[35],
                                "ti_auxtime10": row[36],
                            }
                            result = agentHourlyPerformanceCreation(data, request)
                            if result == True:
                                successCount += 1
                            else:
                                messages.error(request, f"Could not create")
                                # break
                        except Exception as e:
                            logger.info(f"|Failed|Upload failed.Exception: {e}")
                            # break
                    else:
                        break
                if count == successCount:
                    messages.success(
                        request, "ALL Agent Hourly Performance Uploaded Successfully"
                    )
                elif successCount == 0:
                    messages.error(
                        request, "Failed to Upload any Agent Hourly Performance"
                    )
                else:
                    messages.warning(
                        request, "Failed to upload some Agent Hourly Performance"
                    )
            except Exception as e:
                logger.info(
                    f"|Failed| Bulk Agent Hourly Performance Upload failed.Exception: {e}"
                )
                messages.error(request, "Bulk Agent Hourly Performance Upload failed")
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AGENTHOURLYPERFORMANCE_BULK.pageName,
        "formUrl": PageInfoCollection.AGENTHOURLYPERFORMANCE_BULK.urlName,
        "agentHourlyPerformanceActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


################################################################
#   Login Logout Time
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def viewLoginLogoutTime(request):

    templateName = "loginLogoutTime/view.html"
    breadCrumbList = [
        PageInfoCollection.LOGINLOGOUTTIME_VIEW,
    ]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.LOGINLOGOUTTIME_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Login Logout Time",
        },
        "ajaxUrl": PageInfoCollection.SHIFTLEGEND_JSON.urlName,
        "createUrl": PageInfoCollection.SHIFTLEGEND_CREATE.urlName,
        "bulkUrl": PageInfoCollection.LOGINLOGOUTTIME_BULK.urlName,
        "loginLogoutTimeActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def createLoginLogoutTime(request):
    templateName = "avayaCDR/create.html"
    breadCrumbList = [
        PageInfoCollection.AVAYACDR_VIEW,
        PageInfoCollection.AVAYACDR_CREATE,
    ]
    if request.method == "POST":
        form = AvayaCDRForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = avayaCDRCreation(data, request)
            if result == True:
                messages.success(request, "Shift Legend created successfully.")
                return redirect(PageInfoCollection.SHIFTLEGEND_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Shift Legend")
        else:
            messages.error(request, "Could not create Shift Legend")
    else:
        form = AvayaCDRForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.AVAYACDR_CREATE.pageName,
        "formUrl": PageInfoCollection.AVAYACDR_CREATE.urlName,
        "avayaCDRActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


# [136, 298, 299, 300, 301, 305, 317, 318, 320, 322, 325, 328, 336, 344, 364, 372, 488, 498, 526, 529, 542, 571, 616, 667, 715, 742, 753, 879, 884, 909, 915, 924, 925, 929, 930, 932, 935, 946, 948, 982, 986, 987, 988, 990, 992, 1004, 1010, 1022, 1023, 1025, 1034, 1035, 1036, 1038, 1165, 1225, 1254, 1347, 1356, 1429, 1430, 1439, 1443, 1568, 1593, 1620, 1638, 1642, 1651, 1871, 2142, 2147, 2198, 2199, 2215, 2227, 2266]
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1)
def createBulkLoginLogoutTime(request):
    templateName = "loginLogoutTime/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.LOGINLOGOUTTIME_VIEW,
        PageInfoCollection.LOGINLOGOUTTIME_BULK,
    ]
    if request.method == "POST":
        sheetName = request.POST.get("sheetName").strip()
        if not sheetName:
            messages.error(request, "Sheet name is empty")
        else:
            try:
                excel_file = request.FILES["excel_file"]
                wb = load_workbook(excel_file, data_only=True)
                ws = wb[sheetName]
                count = (
                    len(
                        [
                            row
                            for row in ws
                            if not all([cell.value == None for cell in row])
                        ]
                    )
                    - 1
                )
                successCount = 0
                failureList = []
                for index, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=2
                ):
                    logger.info(f"Row {index}: {row}")
                    if row[0] is not None:
                        try:
                            employee = Employee.objects.get(avaya_id=row[0])
                            data = {
                                "employee": employee,
                                "login_date": row[1],
                                "login_time": row[2],
                                "logout_date": row[3],
                                "logout_time": row[4],
                            }
                            result = loginLogoutTimeCreation(data, request)
                            if result == True:
                                successCount += 1
                            else:
                                messages.error(request, f"Could not create")
                                failureList.append(index)
                        except Exception as e:
                            logger.info(f"|Failed|Upload failed.Exception: {e}")
                    else:
                        break
                if count == successCount:
                    messages.success(
                        request, "ALL Login Logout Time Uploaded Successfully"
                    )
                elif successCount == 0:
                    messages.error(request, "Failed to Upload any Login Logout Time")
                else:
                    messages.warning(request, "Failed to upload some Login Logout Time")
                logger.info(f"|Failure List|{failureList}")
            except Exception as e:
                logger.info(
                    f"|Failed| Bulk Login Logout Time Upload failed.Exception: {e}"
                )
                messages.error(request, "Bulk Login Logout Time Upload failed")
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.LOGINLOGOUTTIME_BULK.pageName,
        "formUrl": PageInfoCollection.LOGINLOGOUTTIME_BULK.urlName,
        "loginLogoutTimeActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


################################################################
#   Reporting One
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1, GroupEnum.supervisor)
def viewReportingOne(request):

    templateName = "reportingOne/view.html"
    breadCrumbList = [
        PageInfoCollection.REPORTINGONE_VIEW,
    ]
    # if request.method == "POST":
    #     for key, value in request.POST.items():
    #         #print("Key: %s" % (key))
    #         # #print(f'Key: {key}') in Python >= 3.7
    #         #print("Value %s" % (value))
    #         # #print(f'Value: {value}') in Python >= 3.7
    #     date = request.POST["search_date"]
    #     skill = Skill.objects.get(id=request.POST["skillSelect"])
    #     agentHourlyPerformance = AgentHourlyPerformance.objects.filter(
    #         date=date, skill=skill
    #     )

    #     tableData = [
    #         {"interval": "0", "hour": "0-1"},
    #         {"interval": "1", "hour": "1-2"},
    #         {"interval": "2", "hour": "2-3"},
    #         {"interval": "3", "hour": "3-4"},
    #         {"interval": "4", "hour": "4-5"},
    #         {"interval": "5", "hour": "5-6"},
    #         {"interval": "6", "hour": "6-7"},
    #         {"interval": "7", "hour": "7-8"},
    #         {"interval": "8", "hour": "8-9"},
    #         {"interval": "9", "hour": "9-10"},
    #         {"interval": "10", "hour": "10-11"},
    #         {"interval": "11", "hour": "11-12"},
    #         {"interval": "12", "hour": "12-13"},
    #         {"interval": "13", "hour": "13-14"},
    #         {"interval": "14", "hour": "14-15"},
    #         {"interval": "15", "hour": "15-16"},
    #         {"interval": "16", "hour": "16-17"},
    #         {"interval": "17", "hour": "17-18"},
    #         {"interval": "18", "hour": "18-19"},
    #         {"interval": "19", "hour": "19-20"},
    #         {"interval": "20", "hour": "20-21"},
    #         {"interval": "21", "hour": "21-22"},
    #         {"interval": "22", "hour": "22-23"},
    #         {"interval": "23", "hour": "23-24"},
    #     ]
    #     if skill.name == "premium" or skill.name == "medium":
    #         process = Process.objects.get(name="gp")
    #         lob = LOB.objects.get(name="premium" if skill.name == "premium" else "mass")
    #         forecasts = Forecast.objects.filter(
    #             date=date, process=process, lob=lob
    #         ).order_by("interval")
    #         # Create a mapping of intervals to corresponding forecasts and required_hc
    #         forecast_mapping = {
    #             f.interval: (f.forecast, f.required_hc) for f in forecasts
    #         }
    #         # Update the tableData list with forecast and required_hc

    #     avayaCDRMapping = getAvayaCDRMapping(date=date, skill=skill)
    #     rosterCount = RosterCount.objects.filter(start_date=date)
    #     # Define a list of hour columns
    #     hourColumns = [f"hour_{hour:02d}" for hour in range(24)]
    #     # Create a dictionary to store the total counts for each hour
    #     plannedHeadCount = {}
    #     # Calculate the sum for each hour column individually
    #     for index, hourColumn in enumerate(hourColumns):
    #         headCount = rosterCount.aggregate(total_count=Sum(F(hourColumn)))[
    #             "total_count"
    #         ]
    #         plannedHeadCount[index] = headCount

    #     agentHourlyPerformanceMapping = getAgentHourlyPerformanceMappingWithRoster(
    #         date=date, skill=skill
    #     )

    #     for item in tableData:
    #         interval = int(item["interval"])
    #         shiftCount = 0
    #         isAbsent = 0
    #         # Update fields from AgentHourlyPerformance model
    #         if interval in agentHourlyPerformanceMapping:
    #             performance_data = agentHourlyPerformanceMapping[interval]
    #             shiftCount = performance_data["shiftCount"]
    #             isAbsent = performance_data["absent"]
    #             item["actualHeadCount"] = shiftCount - isAbsent

    #         else:
    #             # Set default values if 'hour' not found in AgentHourlyPerformance
    #             item["absent"] = 0

    #         if interval in forecast_mapping:
    #             forecast, required_hc = forecast_mapping[interval]
    #             item["forecast"] = forecast
    #             item["required_hc"] = required_hc
    #         else:
    #             item["forecast"] = ""  # Set default value if not found
    #             item["required_hc"] = ""  # Set default value if not found
    #         if interval in avayaCDRMapping:
    #             offeredCalls = avayaCDRMapping[interval]
    #             item["offeredCalls"] = offeredCalls

    #         else:
    #             item["offeredCalls"] = ""  # Set default value if not found
    #         # item["plannedHeadCount"] = plannedHeadCount[interval]
    #         item["plannedHeadCount"] = shiftCount
    #         item["plannedHeadCountGap"] = item["required_hc"] - shiftCount
    #         item["actualHeadCountGap"] = item["required_hc"] - item["actualHeadCount"]
    #     # for item in tableData:
    #     #     #print(item)
    #     json_string = json.dumps(tableData)
    #     #print(json_string)
    # else:
    #     tableData = None
    skills = Skill.objects.filter(Q(name="premium") | Q(name="medium"))
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.REPORTINGONE_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Login Logout Time",
        },
        "skills": skills,
        # "tableData": tableData,
        "ajaxUrl": PageInfoCollection.REPORTINGONE_JSON.urlName,
        "formUrl": PageInfoCollection.REPORTINGONE_VIEW.urlName,
        "bulkUrl": PageInfoCollection.LOGINLOGOUTTIME_BULK.urlName,
        "reportingOneActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@method_decorator(
    csrf_exempt, name="dispatch"
)  # Decorator to exempt CSRF validation (for testing purposes)
class viewReportingOneListJson(View):
    def post(self, request, *args, **kwargs):
        # Get the date and skill from the POST request
        date = request.POST.get("date")
        skillID = request.POST.get("skill")
        intervalType = request.POST.get("intervalType")
        monthString = request.POST.get("month")

        logger.info(
            "Received POST request with date: %s, skill ID: %s, intervalType: %s, month: %s",
            date,
            skillID,
            intervalType,
            monthString,
        )

        try:
            skill = Skill.objects.get(id=skillID)
            logger.info("Retrieved skill: %s", skill.name)
        except Skill.DoesNotExist:
            logger.error("Skill with ID %s does not exist", skillID)
            return JsonResponse({"error": "Skill not found"}, status=404)

        tableData = []
        if intervalType == "date":
            tableData = generate_hour_table_data()
            logger.info("Generated hourly table data for date %s", date)
        else:
            year, month = map(int, monthString.split("-"))
            tableData = generate_date_table_data(month, year)
            logger.info(
                "Generated date table data for month: %d, year: %d", month, year
            )

        forecast_mapping = {}
        if skill.name in ["premium", "medium"]:
            try:
                process = Process.objects.get(name="gp")
                lob = LOB.objects.get(
                    name="premium" if skill.name == "premium" else "mass"
                )
                logger.info("Retrieved process: %s, lob: %s", process.name, lob.name)
            except (Process.DoesNotExist, LOB.DoesNotExist) as e:
                logger.error("Process or LOB not found: %s", e)
                return JsonResponse({"error": "Process or LOB not found"}, status=404)

        # Create forecast mapping
        logger.info("Creating Forecast Mapping")
        if intervalType == "date":
            forecasts = Forecast.objects.filter(
                date=date, process=process, lob=lob
            ).order_by("interval")
            forecast_mapping = {
                f.interval: (f.forecast, f.required_hc) for f in forecasts
            }
            logger.info("Forecast mapping by interval created for date: %s", date)
        else:
            forecasts = (
                Forecast.objects.filter(date__month=month, process=process, lob=lob)
                .values("date")
                .annotate(forecast=Sum("forecast"), required_hc=Sum("required_hc"))
            )
            for f in forecasts:
                formatted_date = f["date"].strftime("%Y-%m-%d")
                forecast_mapping[formatted_date] = (f["forecast"], f["required_hc"])
            logger.info(
                "Forecast mapping by date created for month: %s, year: %s",
                month,
                year,
            )

        if intervalType == "date":
            avayaCDRMapping = getAvayaCDRMappingByDate(date=date, skill=skill)
            agentHourlyPerformanceMapping = getAgentHourlyPerformanceMappingWithRoster(
                date=date, skill=skill
            )
            logger.info(
                "Retrieved Avaya CDR and agent performance mapping for date: %s", date
            )
        else:
            avayaCDRMapping = getAvayaCDRMappingByMonth(
                month=month, year=year, skill=skill
            )
            agentHourlyPerformanceMapping = (
                getAgentHourlyPerformanceMappingWithRosterByMonthAI(
                    month=month, year=year, skill=skill
                )
            )
            logger.info(
                "Retrieved Avaya CDR and agent performance mapping for month: %s, year: %s",
                month,
                year,
            )

        # Process table data based on interval type
        try:
            for item in tableData:
                interval = (
                    int(item.get("interval", 0))
                    if intervalType == "date"
                    else item["date"]
                )
                shiftCount = 0
                isAbsent = 0

                # Update fields from AgentHourlyPerformance model
                if interval in agentHourlyPerformanceMapping:
                    performance_data = agentHourlyPerformanceMapping[interval]
                    shiftCount = performance_data["shiftCount"]
                    isAbsent = performance_data["absent"]
                    item["actualHeadCount"] = shiftCount - isAbsent
                    item["plannedHeadCount"] = shiftCount
                else:
                    item["actualHeadCount"] = ""
                    item["plannedHeadCount"] = ""

                # Update fields from forecast_mapping
                if interval in forecast_mapping:
                    forecast, required_hc = forecast_mapping[interval]
                    item["forecast"] = forecast
                    item["required_hc"] = required_hc
                    item["plannedHeadCountGap"] = (
                        item["plannedHeadCount"] - item["required_hc"]
                    )
                    item["actualHeadCountGap"] = (
                        item["actualHeadCount"] - item["plannedHeadCount"]
                    )
                else:
                    item["forecast"] = ""
                    item["required_hc"] = ""
                    item["plannedHeadCountGap"] = ""
                    item["actualHeadCountGap"] = ""

                # Update fields from avayaCDRMapping
                if interval in avayaCDRMapping:
                    avaya_data = avayaCDRMapping[interval]
                    item["offeredCalls"] = avaya_data["offeredCalls"]
                    item["answeredCalls"] = avaya_data["answeredCalls"]
                    item["fd"] = format(
                        item["offeredCalls"] / (item["forecast"] or 1), ".2f"
                    )
                    item["aht"] = avaya_data["aht"]
                else:
                    item["offeredCalls"] = ""
                    item["answeredCalls"] = ""
                    item["fd"] = 0
                    item["aht"] = 0

            logger.info("Processed table data with forecast and performance mappings")
        except Exception as e:
            logger.error("Error processing table data: %s", e)
            return JsonResponse({"error": "Error processing data"}, status=500)

        return JsonResponse(tableData, safe=False)


################################################################
#   Reporting Two
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1, GroupEnum.supervisor)
def viewReportingTwo(request):

    templateName = "reportingTwo/view.html"
    breadCrumbList = [
        PageInfoCollection.REPORTINGTWO_VIEW,
    ]
    skills = Skill.objects.filter(Q(name="premium") | Q(name="medium"))
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.REPORTINGTWO_VIEW.pageName,
        "skills": skills,
        "ajaxUrl": PageInfoCollection.REPORTINGTWO_JSON.urlName,
        "formUrl": PageInfoCollection.REPORTINGTWO_VIEW.urlName,
        "reportingTwoActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


@method_decorator(
    csrf_exempt, name="dispatch"
)  # Decorator to exempt CSRF validation (for testing purposes)
class viewReportingTwoListJson(View):
    def post(self, request, *args, **kwargs):
        # Get the date and skill from the POST request
        date = request.POST.get("date")
        skill = request.POST.get("skill")

        # Your table data
        table_data = [
            {"interval": "0", "hour": "0-1"},
            {"interval": "1", "hour": "1-2"},
            {"interval": "2", "hour": "2-3"},
            {"interval": "3", "hour": "3-4"},
            {"interval": "4", "hour": "4-5"},
            {"interval": "5", "hour": "5-6"},
            {"interval": "6", "hour": "6-7"},
            {"interval": "7", "hour": "7-8"},
            {"interval": "8", "hour": "8-9"},
            {"interval": "9", "hour": "9-10"},
            {"interval": "10", "hour": "10-11"},
            {"interval": "11", "hour": "11-12"},
            {"interval": "12", "hour": "12-13"},
            {"interval": "13", "hour": "13-14"},
            {"interval": "14", "hour": "14-15"},
            {"interval": "15", "hour": "15-16"},
            {"interval": "16", "hour": "16-17"},
            {"interval": "17", "hour": "17-18"},
            {"interval": "18", "hour": "18-19"},
            {"interval": "19", "hour": "19-20"},
            {"interval": "20", "hour": "20-21"},
            {"interval": "21", "hour": "21-22"},
            {"interval": "22", "hour": "22-23"},
            {"interval": "23", "hour": "23-24"},
        ]

        # Your agent hourly performance mapping logic
        agent_hourly_performance_mapping, field_names = (
            getAgentHourlyPerformanceMapping(date=date, skill=skill)
        )

        # Update table data with agent hourly performance mapping
        for item in table_data:
            interval = int(item["interval"])
            if interval in agent_hourly_performance_mapping:
                performance_data = agent_hourly_performance_mapping[interval]
                for key in performance_data:
                    if key != "hour":
                        item[key] = performance_data[key]
            else:
                for key in field_names:
                    if key != "hour":
                        item[key] = 0
        sorted_table_data = sorted(table_data, key=lambda x: int(x["interval"]))
        return JsonResponse(sorted_table_data, safe=False)


################################################################
#   Reporting Three
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.mis_group_1, GroupEnum.supervisor)
def viewReportingThree(request):

    templateName = "reportingThree/view.html"
    breadCrumbList = [
        PageInfoCollection.REPORTINGTHREE_VIEW,
    ]
    tableData = None
    if request.method == "POST":
        date = request.POST["search_date"]
        skill = Skill.objects.get(id=request.POST["skillSelect"])
        tableData = [
            {"interval": "0", "hour": "0-1"},
            {"interval": "1", "hour": "1-2"},
            {"interval": "2", "hour": "2-3"},
            {"interval": "3", "hour": "3-4"},
            {"interval": "4", "hour": "4-5"},
            {"interval": "5", "hour": "5-6"},
            {"interval": "6", "hour": "6-7"},
            {"interval": "7", "hour": "7-8"},
            {"interval": "8", "hour": "8-9"},
            {"interval": "9", "hour": "9-10"},
            {"interval": "10", "hour": "10-11"},
            {"interval": "11", "hour": "11-12"},
            {"interval": "12", "hour": "12-13"},
            {"interval": "13", "hour": "13-14"},
            {"interval": "14", "hour": "14-15"},
            {"interval": "15", "hour": "15-16"},
            {"interval": "16", "hour": "16-17"},
            {"interval": "17", "hour": "17-18"},
            {"interval": "18", "hour": "18-19"},
            {"interval": "19", "hour": "19-20"},
            {"interval": "20", "hour": "20-21"},
            {"interval": "21", "hour": "21-22"},
            {"interval": "22", "hour": "22-23"},
            {"interval": "23", "hour": "23-24"},
        ]

        agentHourlyPerformanceMapping, fieldNames = getAgentHourlyPerformanceMapping(
            date=date, skill=skill
        )

        for item in tableData:
            interval = int(item["interval"])
            # Update fields from AgentHourlyPerformance model
            if interval in agentHourlyPerformanceMapping:
                performance_data = agentHourlyPerformanceMapping[interval]
                for key in performance_data:
                    if key != "hour":  # Exclude 'hour' from the update
                        item[key] = performance_data[key]
            else:
                # Set default values if 'hour' not found in AgentHourlyPerformance
                for (
                    key
                ) in (
                    fieldNames
                ):  # Adjust field_names according to your actual field names
                    if key != "hour":
                        item[key] = 0  # Set default value if not found

    skills = Skill.objects.filter(Q(name="premium") | Q(name="medium"))
    process = Process.objects.get(name="gp")
    employees = Employee.objects.filter(process=process)
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.REPORTINGTHREE_VIEW.pageName,
        "employees": employees,
        "skills": skills,
        "tableData": tableData,
        "ajaxUrl": PageInfoCollection.REPORTINGTHREE_JSON.urlName,
        "formUrl": PageInfoCollection.REPORTINGTWO_VIEW.urlName,
        "bulkUrl": PageInfoCollection.LOGINLOGOUTTIME_BULK.urlName,
        "reportingThreeActive": "active",
        "reportingActive": "active open",
    }
    return render(request, templateName, context)


class viewReportingThreeListJson(BaseDatatableView):
    model = AgentHourlyPerformance

    # Define the columns you want to display
    columns = [
        "date",
        "employee__user__employee_id",
        "employee__user__name",
        "staffed_time",
        "ready_time",
        "short_break",
        "lunch_break",
        "training",
        "meeting",
        "cfs_meeting",
        "one_to_one",
        "outbound_callback",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "date",
        "employee__user__employee_id",
        "employee__user__name",
        "staffed_time",
        "ready_time",
        "short_break",
        "lunch_break",
        "training",
        "meeting",
        "cfs_meeting",
        "one_to_one",
        "outbound_callback",
    ]

    def get_initial_queryset(self):
        # Start timing the execution
        start_time = time.time()
        queryset = None
        if self.request.user.is_WFM() or self.request.user.is_Supervisor():
            # Log the start of the query processing
            logger.info("Starting to fetch initial queryset with filters")
            try:
                # Access query parameters from the orequest
                search_employee = int(self.request.GET.get("search_employee"))
                search_skill = int(self.request.GET.get("search_skill"))
                search_date = self.request.GET.get("search_date")
                if search_date:
                    # Log the received parameters
                    logger.info(
                        f"Received parameters - Employee: {search_employee}, Skill: {search_skill}, Date: {search_date}"
                    )

                    # Start with the base queryset
                    queryset = AgentHourlyPerformance.objects.filter(date=search_date)
                    logger.info(f"Applied filter on date: {search_date}")
                    # Apply filters based on the parameters provided
                    if search_employee != 0:
                        queryset = queryset.filter(employee__id=search_employee)
                        logger.info(f"Applied filter on employee: {search_employee}")
                    if search_skill != 0:
                        queryset = queryset.filter(skill__id=search_skill)
                        logger.info(f"Applied filter on skill: {search_skill}")
                    # Log the success of the queryset filtering
                    logger.info("Successfully applied all filters to queryset")
                else:
                    queryset = AgentHourlyPerformance.objects.all()
                    annotated_qs = queryset.values(
                        "date", "employee__user__employee_id", "employee__user__name"
                    ).annotate(
                        total_staffed_time=Sum("staffed_time"),
                        total_ready_time=Sum("ready_time"),
                        total_short_break=Sum("short_break"),
                        total_lunch_break=Sum("lunch_break"),
                        total_training=Sum("training"),
                        total_meeting=Sum("meeting"),
                        total_cfs_meeting=Sum("cfs_meeting"),
                        total_one_to_one=Sum("one_to_one"),
                        total_outbound_callback=Sum("outbound_callback"),
                    )
                    return annotated_qs
            except ValidationError as ve:
                # Log any validation errors
                logger.error(f"Validation error while filtering queryset: {ve}")
                raise
            except Exception as e:
                # Log unexpected exceptions
                logger.error(f"Unexpected error in get_initial_queryset: {e}")
                raise
        else:
            queryset = AgentHourlyPerformance.objects.all()
        # Calculate and log the execution time
        execution_time = time.time() - start_time
        logger.info(f"get_initial_queryset executed in {execution_time:.4f} seconds")
        return queryset

    def filter_queryset(self, qs):
        # print(f"employee:{self.request.GET.get('search_employee')}",f"skill:{self.request.GET.get('search_skill')}",f"date:{self.request.GET.get('search_date')}",)

        return qs

    def prepare_results(self, qs):
        data = []

        # Annotate the queryset to get the sum of fields GroupEnumed by date and employee
        # annotated_qs = qs.values(
        #     "date", "employee__user__employee_id", "employee__user__name"
        # ).annotate(
        #     total_staffed_time=Sum("staffed_time"),
        #     total_ready_time=Sum("ready_time"),
        #     total_short_break=Sum("short_break"),
        #     total_lunch_break=Sum("lunch_break"),
        #     total_training=Sum("training"),
        #     total_meeting=Sum("meeting"),
        #     total_cfs_meeting=Sum("cfs_meeting"),
        #     total_one_to_one=Sum("one_to_one"),
        #     total_outbound_callback=Sum("outbound_callback"),
        # )

        # for item in qs:
        #     # Fetch the related field and use it directly in the data dictionary
        #     row = {
        #         "date": item.date.strftime("%d-%m-%Y"),
        #         "employee__user__employee_id": item.employee.user.employee_id,
        #         "employee__user__name": item.employee.user.name,
        #         "staffed_time": item.staffed_time,
        #         "ready_time": item.ready_time,
        #         "short_break": item.short_break,
        #         "lunch_break": item.lunch_break,
        #         "training": item.training,
        #         "meeting": item.meeting,
        #         "cfs_meeting": item.cfs_meeting,
        #         "one_to_one": item.one_to_one,
        #         "outbound_callback": item.outbound_callback,
        #     }
        #     data.append(row)

        # for item in annotated_qs:
        #     # Fetch the related field and use it directly in the data dictionary
        #     row = {
        #         "date": item["date"].strftime("%d-%m-%Y"),
        #         "employee__user__employee_id": item["employee__user__employee_id"],
        #         "employee__user__name": item["employee__user__name"],
        #         "staffed_time": item["total_staffed_time"],
        #         "ready_time": item["total_ready_time"],
        #         "short_break": item["total_short_break"],
        #         "lunch_break": item["total_lunch_break"],
        #         "training": item["total_training"],
        #         "meeting": item["total_meeting"],
        #         "cfs_meeting": item["total_cfs_meeting"],
        #         "one_to_one": item["total_one_to_one"],
        #         "outbound_callback": item["total_outbound_callback"],
        #     }
        #     data.append(row)

        for item in qs:
            row = {
                "date": item["date"].strftime("%d-%b-%y"),
                "employee__user__employee_id": item["employee__user__employee_id"],
                "employee__user__name": item["employee__user__name"],
                "staffed_time": item["total_staffed_time"],
                "ready_time": item["total_ready_time"],
                "short_break": item["total_short_break"],
                "lunch_break": item["total_lunch_break"],
                "training": item["total_training"],
                "meeting": item["total_meeting"],
                "cfs_meeting": item["total_cfs_meeting"],
                "one_to_one": item["total_one_to_one"],
                "outbound_callback": item["total_outbound_callback"],
            }
            data.append(row)

        return data

    def render_column(self, row, column):
        return super(viewReportingThreeListJson, self).render_column(row, column)
