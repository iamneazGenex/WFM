from colorama import Fore
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import load_workbook
from .models import Roster, WorkRule, RosterSeatCount
from accounts.models import *
from .forms import *
from django.contrib import messages
from django_datatables_view.base_datatable_view import BaseDatatableView
import logging
from django.contrib import messages
from django.db.models import Q
from rms.global_utilities import *
from .utils import *
from django.utils.html import format_html
from rms.page_info_collection import PageInfoCollection
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from rms.constants import GroupEnum
from rms.decorators import check_user_able_to_see_page
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime, time
from accounts.utils import employeeCreation
from collections import defaultdict
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.models import Group

logger = logging.getLogger(__name__)


# Create your views here.
def import_from_excel(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb["Raw Roster"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Product.objects.create(name=name, price=price, quantity=quantity)
            if row[1] == 20304556 or row[1] == 20305011:
                if row[12] != "OFF":
                    employee = CustomUser.objects.get(employeeID=row[1])
                    roster = Roster.objects.create(
                        employee=employee,
                        startDate=row[11],
                        startTime=row[12],
                        endDate=row[13],
                        endTime=row[14],
                    )
                    roster.save()
            # #print(row[1] = )

        return render(request, "import_success.html")

    return render(request, "roster/import_form.html")


################################################################
#   Roster Management
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewRosterManagement(request):
    templateName = "rosterManagement/view.html"
    breadCrumbList = [PageInfoCollection.ROSTERMANAGEMENT]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTERMANAGEMENT.pageName,
        "ajaxUrl": PageInfoCollection.ROSTERMANAGEMENT.urlName,
        "createUrl": PageInfoCollection.WORKRULE_CREATE.urlName,
        "editUrl": PageInfoCollection.WORKRULE_EDIT.urlName,
        "rosterManagementActive": "active open",
    }
    return render(request, templateName, context)


################################################################
#   Roster
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.supervisor, GroupEnum.employee)
def viewRoster(request):
    templateName = "roster/view.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTER_VIEW,
    ]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTER_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "roster",
        },
        "ajaxUrl": PageInfoCollection.ROSTER_JSON.urlName,
        "createUrl": PageInfoCollection.ROSTER_CREATE.urlName,
        "bulkUrl": PageInfoCollection.ROSTER_BULK.urlName,
        "rosterManagementActive": "active open",
        "rosterActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createRoster(request):
    templateName = "roster/create.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTER_VIEW,
        PageInfoCollection.ROSTER_CREATE,
    ]
    if request.method == "POST":
        form = RosterForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = rosterCreation(data, request)
            if result == True:
                messages.success(request, "Roster created successfully.")
                return redirect(PageInfoCollection.ROSTER_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Roster")
        else:
            messages.error(request, "Could not create Roster")
    else:
        form = RosterForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTER_VIEW.pageName,
        "formUrl": PageInfoCollection.ROSTER_CREATE.urlName,
        "bulkUrl": PageInfoCollection.ROSTER_BULK.urlName,
        "rosterManagementActive": "active open",
        "rosterActive": "active",
    }
    return render(request, templateName, context)


class viewRosterJson(BaseDatatableView):
    model = Roster

    # Define the columns you want to display
    columns = [
        "employee__site__name",
        "employee__user__employee_id",
        "employee__user__system_id",
        "employee__user__name",
        "employee__process__name",
        "employee__lob__name",
        "employee__work_role__name",
        "employee__supervisor_1__user__name",
        "employee__supervisor_2__user__name",
        "employee__gender",
        "employee__pick_drop_location",
        "start_date",
        "start_time",
        "end_date",
        "end_time",
        "created_by",
        "created_At",
        "updated_by",
        "updated_At",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "employee__site__name",
        "employee__user__employee_id",
        "employee__user__system_id",
        "employee__user__name",
        "employee__process__name",
        "employee__lob__name",
        "employee__work_role__name",
        "employee__supervisor_1__user__name",
        "employee__supervisor_2__user__name",
        "employee__gender",
        "employee__pick_drop_location",
        "start_date",
        "start_time",
        "end_date",
        "end_time",
        "created_by",
        "created_At",
        "updated_by",
        "updated_At",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return (
                Roster.objects.all()
                .order_by("-created_At")
                .select_related("shiftLegend")
            )
        elif self.request.user.is_Employee():
            employee = getEmployee(self.request.user.id)
            if employee.is_supervisor():
                supervised_employees = Employee.objects.filter(
                    Q(supervisor_1=employee) | Q(supervisor_2=employee)
                )
                return Roster.objects.filter(
                    Q(employee__in=supervised_employees)
                ).order_by("-created_At")
            else:
                return Roster.objects.filter(
                    Q(employee=employee.id) | Q(employee__lob=employee.lob)
                ).order_by("-created_At")
        else:
            return Roster.objects.none()

    def filter_queryset(self, qs):
        # Handle search parameter from DataTables
        search_value = self.request.GET.get("search[value]", None)

        if search_value:
            # Define the fields you want to search on
            search_fields = [
                "employee__site__name",
                "employee__user__name",
                "employee__process__name",
                "employee__lob__name",
                "employee__work_role__name",
                "employee__supervisor_1__user__name",
                "employee__supervisor_2__user__name",
                "employee__pick_drop_location",
                "start_date",
                "start_time",
                "end_date",
                "end_time",
                "created_by__name",
                "created_At",
                "updated_by__name",
                "updated_At",
            ]

            # Create a Q object to dynamically construct the filter conditions
            search_filter = Q()
            for field in search_fields:
                search_filter |= Q(**{f"{field}__icontains": search_value})

            # Apply the search filter to the queryset
            qs = qs.filter(search_filter)

        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "employee__site__name": item.employee.site.name.title(),
                "employee__user__employee_id": item.employee.user.employee_id,
                "employee__user__system_id": item.employee.user.system_id,
                "employee__user__name": item.employee.user.name,
                "employee__process__name": item.employee.process.name.title(),
                "employee__lob__name": item.employee.lob.name.title(),
                "employee__work_role__name": item.employee.work_role.name.title(),
                "employee__supervisor_1__user__name": (
                    ""
                    if item.employee.supervisor_1 is None
                    else item.employee.supervisor_1.user.name
                ),
                "employee__supervisor_2__user__name": (
                    ""
                    if item.employee.supervisor_2 is None
                    else item.employee.supervisor_2.user.name
                ),
                "employee__gender": item.employee.gender,
                "employee__pick_drop_location": item.employee.pick_drop_location.title(),
                "start_date": (
                    item.start_date.strftime("%d-%m-%Y") if item.start_date else "-"
                ),
                "start_time": (
                    item.start_time.strftime("%H:%M %p") if item.start_time else "-"
                ),
                "end_date": (
                    item.end_date.strftime("%d-%m-%Y") if item.end_date else "-"
                ),
                "end_time": (
                    item.end_time.strftime("%H:%M %p") if item.end_time else "-"
                ),
                "created_by": item.created_by.name,
                "created_at": item.created_At.strftime("%d-%m-%Y"),
                "updated_at": item.updated_At.strftime("%d-%m-%Y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewRosterJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.ROSTER_EDIT.urlName, args=[item.id])
        delete_url = reverse(PageInfoCollection.ROSTER_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}"class="btn btn-success">Edit</a>',
            edit_url,
        )
        delete_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            delete_url,
        )

        return format_html("{} {}", edit_link, delete_link)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createBulkRoster(request):
    templateName = "roster/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTER_VIEW,
        PageInfoCollection.ROSTER_BULK,
    ]
    failedRow = []
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb["Raw Roster"]
        count = (
            len([row for row in ws if not all([cell.value == None for cell in row])])
            - 1
        )
        successCount = 0

        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None:
                logging.info(f"index:{index}")
                failedRow.append({"index": index})
                # #print(row)
                employeeExists = False
                supervisor_Group = Group.objects.get(name="Supervisor")
                supervisors = supervisor_Group.user_set.all()
                supervisor1 = None
                try:
                    employee = Employee.objects.get(user__employee_id=row[1])
                    employeeExists = True
                    # print(employee)
                except Employee.DoesNotExist:
                    # errorMessage = "Employee {employeeName} with Employee ID {employeeID} does not exist".format(
                    #     employeeName=employee.user.name,
                    #     employeeID=employee.user.employee_id,
                    # )
                    # messages.error(request, errorMessage)
                    # break
                    logging.error(f"Employee Does not exist. Creating a new employee.")

                    supervisor1_name = row[8].strip()
                    supervisor1_email = row[9].strip() if row[9] is not None else None
                    if supervisor1_email:
                        supervisor1 = Employee.objects.get(
                            user__email=supervisor1_email, user__in=supervisors
                        )
                        logging.info(f"Supervisor found by email")
                    else:
                        if supervisor1_name:
                            supervisor1 = Employee.objects.get(
                                user__name=supervisor1_name, user__in=supervisors
                            )
                            logging.info(f"Supervisor found by name")
                        else:
                            supervisor1 = None
                            logging.error(f"Supervisor Not found")
                    data = {
                        "name": row[3].strip(),
                        # "email": row[1],
                        "email": (
                            f"{row[3].lower().replace(' ','')}{row[1]}@zohomail.com"
                            if row[4] is None
                            else row[4]
                        ),
                        "employee_id": row[1],
                        "system_id": row[2],
                        "gender": row[11].strip(),
                        "process": row[5].lower().strip(),
                        "site": row[0].lower().strip(),
                        "work_role": row[7].lower().strip(),
                        "lob": (
                            row[6].lower().strip()
                            if isinstance(row[6], str)
                            else row[6]
                        ),
                        "pick_drop_location": row[12].strip(),
                        "supervisor_1": supervisor1,
                        "supervisor_2": None,
                        "password1": "123456",
                    }
                    result = employeeCreation(data, request, True)
                    employeeExists = result
                    if employeeExists is True:
                        employee = Employee.objects.get(user__employee_id=row[1])
                        # print(employee)
                if employeeExists is True:
                    # try:
                    noneShiftNames = [
                        "Dayoff",
                        "LWP",
                        "Mentor",
                        "NTR",
                        "OTR",
                        "PL",
                        "PTR",
                        "Resign",
                        "TR",
                        "Transfer",
                        "Withdraw",
                    ]
                    if row[14] not in noneShiftNames:
                        endTime = time(23, 59)
                        data = {
                            "employee": employee,
                            "start_date": row[13],
                            "start_time": row[14],
                            "end_date": row[15],
                            "end_time": (row[16] if row[16] != endTime else time(0, 0)),
                            "shiftLegend": None,
                            "gender": row[11].strip(),
                            "process": row[5].lower().strip(),
                            "site": row[0].lower().strip(),
                            "work_role": row[7].lower().strip(),
                            "lob": (
                                row[6].lower().strip()
                                if isinstance(row[6], str)
                                else row[6]
                            ),
                            "pick_drop_location": row[12],
                            "supervisor_1": supervisor1,
                            "supervisor_2": None,
                        }
                    else:
                        try:
                            shiftLegend = ShiftLegend.objects.get(shift_name=row[14])
                            data = {
                                "employee": employee,
                                "start_date": row[13],
                                "start_time": None,
                                "end_date": row[15],
                                "end_time": None,
                                "shiftLegend": shiftLegend,
                                "gender": row[11].strip(),
                                "process": row[5].lower().strip(),
                                "site": row[0].lower().strip(),
                                "work_role": row[7].lower().strip(),
                                "lob": (
                                    row[6].lower().strip()
                                    if isinstance(row[6], str)
                                    else row[6]
                                ),
                                "pick_drop_location": row[12].strip(),
                                "supervisor_1": supervisor1,
                                "supervisor_2": None,
                            }
                        except ShiftLegend.DoesNotExist:
                            errorMessage = f"Shift Legend :{row[14]} does not exist"
                            messages.error(request, errorMessage)
                    result = rosterCreation(data, request)
                    # print(result)
                    if result == True:
                        successCount += 1
                        failedRow.remove({"index": index})
                    else:
                        logging.error(f"|Failed| Create A Roster |index:{index}|")
                        break
                # except Exception as e:
                #     logging.error(
                #         f"|Failed| Create A Roster |index:{index}| Exception:{e}"
                #     )
        # print("failedRows:")
        for item in failedRow:
            pass
            # print(item)
        if count == successCount:
            messages.success(request, "Roster Imported Successfully")
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTER_BULK.pageName,
        "formUrl": PageInfoCollection.ROSTER_BULK.urlName,
        "rosterManagementActive": "active open",
        "rosterActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editRoster(request, id):
    templateName = "roster/edit.html"
    roster = Roster.objects.get(id=id)
    # print(id)
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTER_VIEW,
        PageInfoCollection.ROSTER_EDIT,
    ]
    if request.method == "POST":
        form = RosterForm(request.POST, instance=roster)
        if form.is_valid():
            data = form.cleaned_data
            result = rosterModification(id, data)
            if result == True:
                messages.success(request, "Roster Updated successfully.")
                return redirect(PageInfoCollection.ROSTER_VIEW.urlName)
            else:
                messages.error(request, "Failed to update Roster")
        else:
            messages.error(request, "Form is not valid")
    else:
        form = RosterForm(instance=roster)
    context = {
        "form": form,
        "title": "Edit",
        "button": "Update",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTER_EDIT.pageName,
        "formUrl": PageInfoCollection.ROSTER_EDIT.urlName,
        "rosterManagementActive": "active open",
        "rosterActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteRoster(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            roster = Roster.objects.get(id=id)
            roster.delete()
            success = True
        except Roster.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Roster |id:{id}| Exception: Roster does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Roster |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


################################################################
#   Roster Count
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewRosterCount(request):
    templateName = "rosterCount/view.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTERCOUNT_VIEW,
    ]
    sites = Site.objects.all()
    processes = Process.objects.all()
    lobs = LOB.objects.all()
    workRoles = WorkRole.objects.all()
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTERCOUNT_VIEW.pageName,
        "sites": sites,
        "processes": processes,
        "lobs": lobs,
        "workRoles": workRoles,
        "rosterManagementActive": "active open",
        "rosterCountActive": "active",
    }
    return render(request, templateName, context)


class viewRosterCountJson(BaseDatatableView):
    model = RosterCount

    # Define the columns you want to display
    columns = [
        "site__name",
        "process__name",
        "lob__name",
        "work_role__name",
        "start_date",
        "start_time",
        "end_date",
        "end_time",
        "count",
        "hour_00",
        "hour_01",
        "hour_02",
        "hour_03",
        "hour_04",
        "hour_05",
        "hour_06",
        "hour_07",
        "hour_08",
        "hour_09",
        "hour_10",
        "hour_11",
        "hour_12",
        "hour_13",
        "hour_14",
        "hour_15",
        "hour_16",
        "hour_17",
        "hour_18",
        "hour_19",
        "hour_20",
        "hour_21",
        "hour_22",
        "hour_23",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "site__name",
        "process__name",
        "lob__name",
        "work_role__name",
        "start_date",
        "start_time",
        "end_date",
        "end_time",
        "count",
        "hour_00",
        "hour_01",
        "hour_02",
        "hour_03",
        "hour_04",
        "hour_05",
        "hour_06",
        "hour_07",
        "hour_08",
        "hour_09",
        "hour_10",
        "hour_11",
        "hour_12",
        "hour_13",
        "hour_14",
        "hour_15",
        "hour_16",
        "hour_17",
        "hour_18",
        "hour_19",
        "hour_20",
        "hour_21",
        "hour_22",
        "hour_23",
    ]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return RosterCount.objects.all()
        else:
            return RosterCount.objects.none()

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # search = self.request.POST.get("search[value]", None)
        # if search:
        #     # Filter by employee's name
        #     qs = qs.filter(process__name__icontains=search)

        # Apply date range filter if provided in the request parameters
        # start_date_range = self.request.GET.get("start_date_range")
        # end_date_range = self.request.GET.get("end_date_range")
        # #print(f"start: {start_date_range} end: {end_date_range}")
        # if start_date_range and end_date_range:
        #     qs = qs.filter(start_date__range=[start_date_range, end_date_range])
        # return qs

        # Apply date range filter if provided in the request parameters
        search_date = self.request.GET.get("search_date")
        search_site = int(self.request.GET.get("search_site"))
        search_process = int(self.request.GET.get("search_process"))
        search_lob = int(self.request.GET.get("search_lob"))
        search_workRole = int(self.request.GET.get("search_workRole"))
        # print(search_date, search_site, search_process, search_lob, search_workRole)
        if search_site != 0:
            qs = qs.filter(site=search_site)
        if search_process != 0:
            qs = qs.filter(process=search_process)
        if search_lob != 0:
            qs = qs.filter(lob=search_lob)
        if search_workRole != 0:
            qs = qs.filter(workRole=search_workRole)
        if search_date:
            qs = qs.filter(start_date=search_date)

        # print(qs)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "site__name": item.site.name.title(),
                "process__name": item.process.name.title(),
                "lob__name": item.lob.name.title(),
                "work_role__name": item.workRole.name.title(),
                "start_date": item.start_date.strftime("%d-%m-%Y"),
                "start_time": item.start_time.strftime("%H:%M %p"),
                "end_date": item.end_date.strftime("%d-%m-%Y"),
                "end_time": item.end_time.strftime("%H:%M %p"),
                "count": item.count,
                "hour_00": item.hour_00,
                "hour_01": item.hour_01,
                "hour_02": item.hour_02,
                "hour_03": item.hour_03,
                "hour_04": item.hour_04,
                "hour_05": item.hour_05,
                "hour_06": item.hour_06,
                "hour_07": item.hour_07,
                "hour_08": item.hour_08,
                "hour_09": item.hour_09,
                "hour_10": item.hour_10,
                "hour_11": item.hour_11,
                "hour_12": item.hour_12,
                "hour_13": item.hour_13,
                "hour_14": item.hour_14,
                "hour_15": item.hour_15,
                "hour_16": item.hour_16,
                "hour_17": item.hour_17,
                "hour_18": item.hour_18,
                "hour_19": item.hour_19,
                "hour_20": item.hour_20,
                "hour_21": item.hour_21,
                "hour_22": item.hour_22,
                "hour_23": item.hour_23,
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewRosterCountJson, self).render_column(row, column)


################################################################
#   Roster Seat Count
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewRosterSeatCount(request):
    templateName = "rosterSeatCount/view.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTERSEATCOUNT_VIEW,
    ]
    sites = Site.objects.all()
    processes = Process.objects.all()
    lobs = LOB.objects.all()
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTERSEATCOUNT_VIEW.pageName,
        "sites": sites,
        "processes": processes,
        "lobs": lobs,
        "createUrl": PageInfoCollection.ROSTERSEATCOUNT_CREATE.urlName,
        "bulkUrl": PageInfoCollection.ROSTERSEATCOUNT_BULK.urlName,
        "ajaxUrl": PageInfoCollection.ROSTERSEATCOUNT_JSON.urlName,
        "rosterManagementActive": "active open",
        "rosterSeatCountActive": "active",
    }
    return render(request, templateName, context)


class viewRosterSeatCountJson(BaseDatatableView):
    model = RosterSeatCount

    # Define the columns you want to display
    columns = [
        "site__name",
        "process__name",
        "lob__name",
        "total",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = ["site__name", "process__name", "lob__name", "total", ""]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return RosterSeatCount.objects.all()
        else:
            return RosterSeatCount.objects.none()

    def filter_queryset(self, qs):
        search_site = int(self.request.GET.get("search_site"))
        search_process = int(self.request.GET.get("search_process"))
        search_lob = int(self.request.GET.get("search_lob"))

        if search_site != 0:
            qs = qs.filter(site=search_site)
        if search_process != 0:
            qs = qs.filter(process=search_process)
        if search_lob != 0:
            qs = qs.filter(lob=search_lob)

        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "site__name": item.site.name.upper(),
                "process__name": item.process.name.title(),
                "lob__name": item.lob.name.title(),
                "total": item.total,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewRosterSeatCountJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(
            PageInfoCollection.ROSTERSEATCOUNT_EDIT.urlName, args=[item.id]
        )
        delete_url = reverse(
            PageInfoCollection.ROSTERSEATCOUNT_DELETE.urlName, args=[item.id]
        )

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        delete_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            delete_url,
        )

        return format_html("{} {}", edit_link, delete_link)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createRosterSeatCount(request):
    templateName = "rosterSeatCount/create.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTERSEATCOUNT_VIEW,
        PageInfoCollection.ROSTERSEATCOUNT_CREATE,
    ]
    if request.method == "POST":
        form = RosterSeatCountForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = rosterSeatCountCreation(data, request)
            if result == True:
                messages.success(request, "Roster Seat Count created successfully.")
                return redirect(PageInfoCollection.ROSTERSEATCOUNT_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Roster Seat Count")
        else:
            messages.error(request, "Could not create Roster Seat Count")
    else:
        form = RosterSeatCountForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTERSEATCOUNT_CREATE.pageName,
        "formUrl": PageInfoCollection.ROSTERSEATCOUNT_CREATE.urlName,
        "rosterManagementActive": "active open",
        "rosterSeatCountActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createBulkRosterSeatCount(request):
    templateName = "rosterSeatCount/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTERSEATCOUNT_VIEW,
        PageInfoCollection.ROSTERSEATCOUNT_BULK,
    ]
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb["Sheet1"]
        count = (
            len([row for row in ws if not all([cell.value == None for cell in row])])
            - 1
        )
        successCount = 0
        failedList = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                # print(row)
                try:
                    data = {
                        "site": Site.objects.get(name=row[0].lower()),
                        "process": Process.objects.get(name=row[2].lower()),
                        "lob": LOB.objects.get(name=row[3].lower()),
                        "total": row[4],
                    }
                    result = rosterSeatCountCreation(data, request)
                    if result == True:
                        successCount += 1
                    else:
                        failedList.append(row[0])
                except Exception as e:
                    logger.error(
                        f"|Failed| Failed to create Roster Seat Count of Site: {row[0]} | Process: {row[2]} | Lob: {row[3]}.\nException: {e}"
                    )
                    messages.error(request, "Could not create Roster Seat Count")
        if count == successCount:
            messages.success(request, "All Roster Seat Count Created successfully.")
        else:
            messages.error(request, f"Failed : {','.join(failedList)}")
        return redirect(PageInfoCollection.ROSTERSEATCOUNT_VIEW.urlName)
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTERSEATCOUNT_BULK.pageName,
        "formUrl": PageInfoCollection.ROSTERSEATCOUNT_BULK.urlName,
        "rosterManagementActive": "active open",
        "rosterSeatCountActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editRosterSeatCount(request, id):
    templateName = "rosterSeatCount/edit.html"
    instance = RosterSeatCount.objects.get(id=id)
    breadCrumbList = [
        PageInfoCollection.ROSTERMANAGEMENT,
        PageInfoCollection.ROSTERSEATCOUNT_VIEW,
        PageInfoCollection.ROSTERSEATCOUNT_EDIT,
    ]
    if request.method == "POST":
        form = RosterSeatCountForm(request.POST, instance=instance)
        if form.is_valid():
            data = form.cleaned_data
            result = rosterSeatCountEdit(
                id=id,
                data=data,
                updated_by=request.user.id,
            )
            if result == True:
                messages.success(request, "Roster Seat Count Updated successfully.")
                return redirect(PageInfoCollection.ROSTERSEATCOUNT_VIEW.urlName)
            else:
                messages.error(request, "Failed to update Roster Seat Count")
        else:
            messages.error(request, "Form is not valid")
    else:
        form = RosterSeatCountForm(instance=instance)
    context = {
        "form": form,
        "title": "Edit",
        "button": "Update",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.ROSTERSEATCOUNT_EDIT.pageName,
        "formUrl": PageInfoCollection.ROSTERSEATCOUNT_EDIT.urlName,
        "rosterManagementActive": "active open",
        "rosterSeatCountActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteRosterSeatCount(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            roster = RosterSeatCount.objects.get(id=id)
            roster.delete()
            success = True
        except Roster.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Roster Seat Count |id:{id}| Exception: Roster Seat Count does not exist."
            )
        except Exception as e:
            logging.error(
                f"|Failed| Delete A Roster Seat Count |id:{id}| Exception:{e}"
            )
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


################################################################
#   Work RULE
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewWorkRule(request):
    templateName = "workRule/view.html"
    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.WORKRULE_VIEW]
    try:
        workRule = WorkRule.objects.get(id=1)
    except Exception as e:
        workRule = None
        logger.info(f"|Failed| Roster Count Creation failed.Exception: {e}")
        messages.error(request, "WorkRule does not exist.")
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.WORKRULE_VIEW.pageName,
        "workRule": workRule,
        "ajaxUrl": PageInfoCollection.EMPLOYEE_JSON.urlName,
        "createUrl": PageInfoCollection.WORKRULE_CREATE.urlName,
        "editUrl": PageInfoCollection.WORKRULE_EDIT.urlName,
        "settingsActive": "active open",
        "workRuleActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createWorkRule(request):
    template_name = "workRule/create.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.WORKRULE_VIEW,
        PageInfoCollection.WORKRULE_CREATE,
    ]
    if request.method == "POST":
        form = WorkRuleForm(request.POST)
        if form.is_valid():
            try:
                workRule = form.save(commit=False)
                workRule.created_by = request.user
                workRule.save()
                messages.success(request, "WorkRule created successfully.")
                logger.info(f"|Success| WorkRule created successfully")
                return redirect(PageInfoCollection.WORKRULE_VIEW.urlName)
            except Exception as e:
                messages.error(request, "WorkRule Creation failed")
                logger.info(f"|Failed| WorkRule Creation failed.Exception: {e}")
    else:
        form = WorkRuleForm()

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.WORKRULE_CREATE.pageName,
        "formUrl": PageInfoCollection.WORKRULE_CREATE.urlName,
        "settingsActive": "active open",
        "workRuleActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editWorkRule(request, id):
    template_name = "workRule/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.WORKRULE_VIEW,
        PageInfoCollection.WORKRULE_EDIT,
    ]
    workRule = WorkRule.objects.get(id=1)
    if request.method == "POST":
        form = WorkRuleForm(request.POST, instance=workRule)
        if form.is_valid():
            try:
                workRule = form.save(commit=False)
                workRule.updated_by = request.user
                workRule.save()
                logger.info(f"|Success| WorkRule updated successfully")
                messages.success(request, "WorkRule updated successfully.")
                return redirect(PageInfoCollection.WORKRULE_VIEW.urlName)
            except Exception as e:
                messages.error(request, "WorkRule Update failed")
                logger.info(f"|Failed| WorkRule Update failed.Exception: {e}")

    else:
        form = WorkRuleForm(instance=workRule)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.WORKRULE_EDIT.pageName,
        "formUrl": PageInfoCollection.WORKRULE_EDIT.urlName,
        "settingsActive": "active open",
        "workRuleActive": "active",
    }
    return render(request, template_name, context)


################################################################
#   Forecasting
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.supervisor, GroupEnum.employee)
def viewForecasting(request):

    date = datetime(2024, 1, 1)
    site = Site.objects.get(name="dhk")
    process = Process.objects.get(name="gp")
    lob = LOB.objects.get(name="mass")
    forecastings = (
        Forecast.objects.filter(date=date, process=process, lob=lob)
        .order_by("interval")
        .values()
    )
    rosterCounts = RosterCount.objects.filter(process=process, lob=lob, start_date=date)
    # Initialize a dictionary to store the sum for each hour
    hour_sum = {f"hour_{hour:02d}": 0 for hour in range(24)}
    for rosterCount in rosterCounts:
        for hour in range(24):
            field_name = f"hour_{hour:02d}"
            hour_sum[field_name] += getattr(rosterCount, field_name)

    # hour_sum_list = list(hour_sum.items())

    for forecast in forecastings:
        interval = forecast["interval"]
        if interval < 10:
            fieldName = f"hour_0{interval}"
        else:
            fieldName = f"hour_{interval}"
        # print(f"Interval : {interval} - Forecast :{forecast['forecast']} - Required Head Count:{forecast['required_hc']} - RosterCount : {hour_sum[fieldName]}")
    templateName = "forecasting/view.html"
    breadCrumbList = [
        PageInfoCollection.FORECASTING_VIEW,
    ]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.FORECASTING_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Forecast",
        },
        "ajaxUrl": PageInfoCollection.FORECASTING_JSON.urlName,
        "createUrl": PageInfoCollection.FORECASTING_CREATE.urlName,
        "bulkUrl": PageInfoCollection.FORECASTING_BULK.urlName,
        "forecastingActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createForecasting(request):
    templateName = "forecasting/create.html"
    breadCrumbList = [
        PageInfoCollection.FORECASTING_VIEW,
        PageInfoCollection.FORECASTING_CREATE,
    ]
    if request.method == "POST":
        form = ForecastingForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = forecastingCreation(data, request)
            if result == True:
                messages.success(request, "Forecasting created successfully.")
                return redirect(PageInfoCollection.FORECASTING_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Forecasting")
        else:
            messages.error(request, "Could not create Forecasting")
    else:
        form = ForecastingForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.FORECASTING_CREATE.pageName,
        "formUrl": PageInfoCollection.FORECASTING_CREATE.urlName,
        "forecastingActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createBulkForcasting(request):
    templateName = "forecasting/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.FORECASTING_VIEW,
        PageInfoCollection.FORECASTING_BULK,
    ]
    # print("here")
    if request.method == "POST":
        try:
            excel_file = request.FILES["excel_file"]
            wb = load_workbook(excel_file, data_only=True)
            ws = wb["Sheet1"]
            count = (
                len(
                    [row for row in ws if not all([cell.value == None for cell in row])]
                )
                - 1
            )
            successCount = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    # print(row)
                    try:
                        process = Process.objects.get(name=row[1].lower())
                        lob = LOB.objects.get(name=row[2].lower())
                        data = {
                            "date": row[0],
                            "process": process,
                            "lob": lob,
                            "interval": row[3],
                            "forecast": row[4],
                            "required_hc": row[5],
                        }
                        result = forecastingCreation(data, request)
                        if result == True:
                            successCount += 1
                        # else:
                        #     break
                    except ObjectDoesNotExist as e:
                        logger.info(
                            f"|Failed| {e.model.__name__} does not exist .Exception: {e}"
                        )
                    except Exception as e:
                        pass
                        # print(f"An error occurred: {e}")
            if count == successCount:
                messages.success(request, "ALL Forecasting Uploaded Successfully")
            elif successCount == 0:
                messages.error(request, "Failed to Upload any Forecasting")
            else:
                messages.warning(request, "Failed to upload some Forecasting")
        except Exception as e:
            logger.info(f"|Failed| Bulk Forecasting Upload failed.Exception: {e}")
            messages.error(request, "Bulk Forecasting Upload failed")
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.FORECASTING_BULK.pageName,
        "formUrl": PageInfoCollection.FORECASTING_BULK.urlName,
        "forecastingActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editForecasting(request, id):
    templateName = "forecasting/edit.html"
    forecastInstance = Forecast.objects.get(id=id)
    breadCrumbList = [
        PageInfoCollection.FORECASTING_VIEW,
        PageInfoCollection.FORECASTING_EDIT,
    ]
    if request.method == "POST":
        form = ForecastingForm(request.POST, instance=forecastInstance)
        if form.is_valid():
            data = form.cleaned_data
            result = forecastingModification(id, data)
            if result == True:
                messages.success(request, "Forecasting Updated successfully.")
                return redirect(PageInfoCollection.FORECASTING_VIEW.urlName)
            else:
                messages.error(request, "Failed to update Forecasting")
        else:
            messages.error(request, "Form is not valid")
    else:
        form = ForecastingForm(instance=forecastInstance)
    context = {
        "form": form,
        "title": "Edit",
        "button": "Update",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.FORECASTING_EDIT.pageName,
        "formUrl": PageInfoCollection.FORECASTING_EDIT.urlName,
        "forecastingActive": "active",
    }
    return render(request, templateName, context)


class viewForecastingListJson(BaseDatatableView):
    model = Forecast

    # Define the columns you want to display
    columns = [
        "date",
        "process__name",
        "lob__name",
        "interval",
        "forecast",
        "required_hc",
        "created_by",
        "created_At",
        "updated_by",
        "updated_At",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "date",
        "process__name",
        "lob__name",
        "interval",
        "forecast",
        "required_hc",
        "created_by",
        "created_At",
        "updated_by",
        "updated_At",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return Forecast.objects.all().order_by("-created_At")
        else:
            return Forecast.objects.none()

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        search_user = self.request.GET.get("search_user", None)
        # print(search_user)
        if search_user:
            qs = qs.filter(employee__user__name=search_user)
        # print(qs)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "date": (item.date.strftime("%d-%m-%Y") if item.date else "-"),
                "process__name": item.process.name.title(),
                "lob__name": item.lob.name.title(),
                "interval": item.interval,
                "forecast": item.forecast,
                "required_hc": item.required_hc,
                "created_by": "" if item.updated_by is None else item.updated_by.name,
                "created_at": item.created_At.strftime("%d-%m-%Y"),
                "updated_at": item.updated_At.strftime("%d-%m-%Y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewForecastingListJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.FORECASTING_EDIT.urlName, args=[item.id])
        delete_url = reverse(
            PageInfoCollection.FORECASTING_DELETE.urlName, args=[item.id]
        )
        edit_link = format_html(
            '<a href="{}"class="btn btn-success">Edit</a>',
            edit_url,
        )
        delete_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            delete_url,
        )

        return format_html("{} {}", edit_link, delete_link)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteForecasting(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            forecastInstance = Forecast.objects.get(id=id)
            forecastInstance.delete()
            success = True
        except Forecast.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Forecast |id:{id}| Exception: Forecast does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Forecast |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


################################################################
#   SHIFT LEGEND
################################################################
@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.supervisor, GroupEnum.employee)
def viewShiftLegend(request):

    templateName = "shiftLegend/view.html"
    breadCrumbList = [
        PageInfoCollection.SHIFTLEGEND_VIEW,
    ]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SHIFTLEGEND_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Shift Legend",
        },
        "ajaxUrl": PageInfoCollection.SHIFTLEGEND_JSON.urlName,
        "createUrl": PageInfoCollection.SHIFTLEGEND_CREATE.urlName,
        "bulkUrl": PageInfoCollection.SHIFTLEGEND_BULK.urlName,
        "shiftLegendActive": "active",
        "settingsActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createShiftLegend(request):
    templateName = "shiftLegend/create.html"
    breadCrumbList = [
        PageInfoCollection.SHIFTLEGEND_VIEW,
        PageInfoCollection.SHIFTLEGEND_CREATE,
    ]
    if request.method == "POST":
        form = ShiftLegendForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = shiftLegendCreation(data, request)
            if result == True:
                messages.success(request, "Shift Legend created successfully.")
                return redirect(PageInfoCollection.SHIFTLEGEND_VIEW.urlName)
            else:
                messages.error(request, "Failed to create Shift Legend")
        else:
            messages.error(request, "Could not create Shift Legend")
    else:
        form = ShiftLegendForm()
    context = {
        "form": form,
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SHIFTLEGEND_CREATE.pageName,
        "formUrl": PageInfoCollection.SHIFTLEGEND_CREATE.urlName,
        "settingsActive": "active open",
        "shiftLegendActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createBulkShiftLegend(request):
    templateName = "shiftLegend/create_bulk.html"
    breadCrumbList = [
        PageInfoCollection.SHIFTLEGEND_VIEW,
        PageInfoCollection.SHIFTLEGEND_BULK,
    ]
    if request.method == "POST":
        sheetName = request.POST.get("sheetName").strip()
        if not sheetName:
            messages.error(request, "Sheet name is empty")
        else:
            try:
                excel_file = request.FILES["excel_file"]
                wb = load_workbook(excel_file, data_only=True)
                ws = wb[sheetName]
                count = (
                    len(
                        [
                            row
                            for row in ws
                            if not all([cell.value == None for cell in row])
                        ]
                    )
                    - 1
                )
                successCount = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    logger.info(f"Row:{count + 1}")
                    if row[0] is not None:
                        try:
                            shiftName = row[0]
                            shiftCount = row[1]
                            shiftStartTime = row[2]
                            shiftEndTime = row[3]

                            data = {
                                "shift_name": shiftName,
                                "shift_start_time": shiftStartTime,
                                "shift_end_time": shiftEndTime,
                                "shift_count": shiftCount,
                            }
                            result = shiftLegendCreation(data, request)
                            if result == True:
                                successCount += 1
                            else:
                                messages.error(request, shiftName)
                        except Exception as e:
                            logger.info(
                                f"|Failed| {shiftStartTime}-{shiftEndTime} Upload failed.Exception: {e}"
                            )
                if count == successCount:
                    messages.success(request, "ALL ShiftLegend Uploaded Successfully")
                elif successCount == 0:
                    messages.error(request, "Failed to Upload any ShiftLegend")
                else:
                    messages.warning(request, "Failed to upload some ShiftLegend")
            except Exception as e:
                logger.info(f"|Failed| Bulk ShiftLegend Upload failed.Exception: {e}")
                messages.error(request, "Bulk ShiftLegend Upload failed")
    context = {
        "title": "Create",
        "button": "Create",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SHIFTLEGEND_BULK.pageName,
        "formUrl": PageInfoCollection.SHIFTLEGEND_BULK.urlName,
        "settingsActive": "active open",
        "shiftLegendActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editShiftLegend(request, id):
    templateName = "shiftLegend/edit.html"
    shiftLegendInstance = ShiftLegend.objects.get(id=id)
    breadCrumbList = [
        PageInfoCollection.SHIFTLEGEND_VIEW,
        PageInfoCollection.SHIFTLEGEND_EDIT,
    ]
    if request.method == "POST":
        form = ShiftLegendForm(request.POST, instance=shiftLegendInstance)
        if form.is_valid():
            data = form.cleaned_data
            result = shiftLegendModification(id, data)
            if result == True:
                messages.success(request, "Shift Legend Updated successfully.")
                return redirect(PageInfoCollection.SHIFTLEGEND_VIEW.urlName)
            else:
                messages.error(request, "Failed to update Shift Legend")
        else:
            messages.error(request, "Form is not valid")
    else:
        form = ShiftLegendForm(instance=shiftLegendInstance)
    context = {
        "form": form,
        "title": "Edit",
        "button": "Update",
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SHIFTLEGEND_EDIT.pageName,
        "formUrl": PageInfoCollection.SHIFTLEGEND_EDIT.urlName,
        "forecastingActive": "active",
    }
    return render(request, templateName, context)


class viewShiftLegendListJson(BaseDatatableView):
    model = ShiftLegend

    # Define the columns you want to display
    columns = [
        "shift_name",
        "shift_count",
        "shift_start_time",
        "shift_end_time",
        "login_start_hour",
        "duty_hour",
        "target_break",
        "target_ready",
        "hour_0",
        "hour_1",
        "hour_2",
        "hour_3",
        "hour_4",
        "hour_5",
        "hour_6",
        "hour_7",
        "hour_8",
        "hour_9",
        "hour_10",
        "hour_11",
        "hour_12",
        "hour_13",
        "hour_14",
        "hour_15",
        "hour_16",
        "hour_17",
        "hour_18",
        "hour_19",
        "hour_20",
        "hour_21",
        "hour_22",
        "hour_23",
        "created_by",
        "updated_by",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "shift_name",
        "shift_count",
        "shift_start_time",
        "shift_end_time",
        "login_start_hour",
        "duty_hour",
        "target_break",
        "target_ready",
        "hour_0",
        "hour_1",
        "hour_2",
        "hour_3",
        "hour_4",
        "hour_5",
        "hour_6",
        "hour_7",
        "hour_8",
        "hour_9",
        "hour_10",
        "hour_11",
        "hour_12",
        "hour_13",
        "hour_14",
        "hour_15",
        "hour_16",
        "hour_17",
        "hour_18",
        "hour_19",
        "hour_20",
        "hour_21",
        "hour_22",
        "hour_23",
        "created_by",
        "updated_by",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return ShiftLegend.objects.all().order_by("-created_At")
        else:
            return ShiftLegend.objects.none()

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        search_user = self.request.GET.get("search_user", None)
        # print(search_user)
        if search_user:
            qs = qs.filter(employee__user__name=search_user)
        # print(qs)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "shift_name": item.shift_name,
                "shift_count": item.shift_count,
                "shift_start_time": (
                    ""
                    if item.shift_start_time is None
                    else item.shift_start_time.strftime("%I:%M %p")
                ),
                "shift_end_time": (
                    ""
                    if item.shift_end_time is None
                    else item.shift_end_time.strftime("%I:%M %p")
                ),
                "login_start_hour": (
                    "" if item.login_start_hour is None else item.login_start_hour
                ),
                "duty_hour": (
                    ""
                    if item.duty_hour is None
                    else item.duty_hour.total_seconds() // 3600
                ),
                "target_break": "" if item.target_break is None else item.target_break,
                "target_ready": "" if item.target_ready is None else item.target_ready,
                "hour_0": (
                    "" if item.hour_0 is False or item.hour_0 is None else item.hour_0
                ),
                "hour_1": (
                    "" if item.hour_1 is False or item.hour_1 is None else item.hour_1
                ),
                "hour_2": (
                    "" if item.hour_2 is False or item.hour_2 is None else item.hour_2
                ),
                "hour_3": (
                    "" if item.hour_3 is False or item.hour_3 is None else item.hour_3
                ),
                "hour_4": (
                    "" if item.hour_4 is False or item.hour_4 is None else item.hour_4
                ),
                "hour_5": (
                    "" if item.hour_5 is False or item.hour_5 is None else item.hour_5
                ),
                "hour_6": (
                    "" if item.hour_6 is False or item.hour_6 is None else item.hour_6
                ),
                "hour_7": (
                    "" if item.hour_7 is False or item.hour_7 is None else item.hour_7
                ),
                "hour_8": (
                    "" if item.hour_8 is False or item.hour_8 is None else item.hour_8
                ),
                "hour_9": (
                    "" if item.hour_9 is False or item.hour_9 is None else item.hour_9
                ),
                "hour_10": (
                    ""
                    if item.hour_10 is False or item.hour_10 is None
                    else item.hour_10
                ),
                "hour_11": (
                    ""
                    if item.hour_11 is False or item.hour_11 is None
                    else item.hour_11
                ),
                "hour_12": (
                    ""
                    if item.hour_12 is False or item.hour_12 is None
                    else item.hour_12
                ),
                "hour_13": (
                    ""
                    if item.hour_13 is False or item.hour_13 is None
                    else item.hour_13
                ),
                "hour_14": (
                    ""
                    if item.hour_14 is False or item.hour_14 is None
                    else item.hour_14
                ),
                "hour_15": (
                    ""
                    if item.hour_15 is False or item.hour_15 is None
                    else item.hour_15
                ),
                "hour_16": (
                    ""
                    if item.hour_16 is False or item.hour_16 is None
                    else item.hour_16
                ),
                "hour_17": (
                    ""
                    if item.hour_17 is False or item.hour_17 is None
                    else item.hour_17
                ),
                "hour_18": (
                    ""
                    if item.hour_18 is False or item.hour_18 is None
                    else item.hour_18
                ),
                "hour_19": (
                    ""
                    if item.hour_19 is False or item.hour_19 is None
                    else item.hour_19
                ),
                "hour_20": (
                    ""
                    if item.hour_20 is False or item.hour_20 is None
                    else item.hour_20
                ),
                "hour_21": (
                    ""
                    if item.hour_21 is False or item.hour_21 is None
                    else item.hour_21
                ),
                "hour_22": (
                    ""
                    if item.hour_22 is False or item.hour_22 is None
                    else item.hour_22
                ),
                "hour_23": (
                    ""
                    if item.hour_23 is False or item.hour_23 is None
                    else item.hour_23
                ),
                "created_by": "" if item.created_by is None else item.created_by.name,
                "created_at": item.created_At.strftime("%d-%m-%Y"),
                "updated_at": item.updated_At.strftime("%d-%m-%Y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewForecastingListJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.SHIFTLEGEND_EDIT.urlName, args=[item.id])
        delete_url = reverse(
            PageInfoCollection.SHIFTLEGEND_DELETE.urlName, args=[item.id]
        )
        edit_link = format_html(
            '<a href="{}"class="btn btn-success">Edit</a>',
            edit_url,
        )
        delete_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            delete_url,
        )

        return format_html("{} {}", edit_link, delete_link)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteShiftLegend(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            shiftLegendInstance = ShiftLegend.objects.get(id=id)
            shiftLegendInstance.delete()
            success = True
        except ShiftLegend.DoesNotExist:
            logging.error(
                f"|Failed| Delete A ShiftLegend |id:{id}| Exception: ShiftLegend does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A ShiftLegend |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


@require_GET
def get_employee_schedule(request, *args, **kwargs):
    """
    Fetches roster entries for an employee within a date range.
    Args:
        id (int): The employee ID for whom to retrieve the schedule.
        start_date (str, optional): Optional start date for filtering. Defaults to today.
        end_date (str, optional): Optional end date for filtering. Defaults to 7 days from today.

    Returns:
        JsonResponse: A JSON response with a list of dictionaries representing roster entries.
    """
    if "id" in kwargs:
        id = kwargs["id"]
    else:
        id = args[0]  # Assuming the first argument is the id
    start_date = kwargs.get("start_date")
    end_date = kwargs.get("end_date")
    if not start_date:
        start_date = datetime.today().date()
    else:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    if not end_date:
        end_date = start_date + timedelta(days=7)
    else:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    try:
        employee = Employee.objects.get(id=id)
        rosters = Roster.objects.filter(
            employee=employee, start_date__gte=start_date, end_date__lte=end_date
        ).values(
            "start_date",
            "end_date",
            "start_time",
            "end_time",
            "shiftLegend__shift_name",
        )
        rosters_list = list(rosters)
        return JsonResponse(rosters_list, safe=False)
    except Employee.DoesNotExist:
        return JsonResponse({"error": "Employee not found"}, status=404)
