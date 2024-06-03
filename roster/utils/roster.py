import logging
from openpyxl import load_workbook
from django.contrib import messages
from django.shortcuts import render
from accounts.models import (
    Employee,
    Process,
    Site,
    WorkRole,
    LOB,
    CustomUser,
)
from ..models import *
from .common import *
from django.contrib.auth.models import Group
from django.shortcuts import get_object_or_404
from datetime import time
from rms.global_utilities import *

logger = logging.getLogger(__name__)


def getSupervisor(name=None, email=None):
    """Get a CustomUser in the Supervisor group by name or email.

    Args:
        name (str, optional): Name of the Supervisor.
        email (str, optional): Email of the Supervisor.

    Returns:
        CustomUser or None: The Supervisor if found, otherwise None.
    """
    supervisor = None
    try:
        supervisorGroup = Group.objects.get(name="Supervisor")
        supervisors = supervisorGroup.user_set.all()
        if email:
            email = email.strip()
            supervisor = get_object_or_404(CustomUser, email=email, id__in=supervisors)
            logging.info(
                f"Supervisor : {supervisor.name}[{supervisor.id}] found by email"
            )
        elif name:
            name = name.strip()
            supervisor = get_object_or_404(CustomUser, name=name, id__in=supervisors)
            logging.info(
                f"Supervisor: {supervisor.name}[{supervisor.id}] found by name"
            )
        else:
            logging.error("No name or email provided to find Supervisor")
    except Group.DoesNotExist:
        logger.error("Supervisor group does not exist.")
    except CustomUser.DoesNotExist:
        logger.error(f"User does not exist")
    except Exception as e:
        logger.error(f"An error occurred while finding the supervisor:{e}")
    return supervisor


def process_roster_file(request, sheet_name, failed_rows):
    try:
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
        total_rows = count_non_empty_rows(ws) - 1
        success_count = 0

        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None:
                logging.info("--------------------------------")
                logging.info(f"Processing row index: {index}")
                failed_rows.append({"index": index})

                if process_roster_row(request, row, index):
                    success_count += 1
                    failed_rows.remove({"index": index})
                else:
                    break

        log_failed_rows(failed_rows)

        if total_rows == success_count:
            logging.info("All Rosters Imported Successfully")
            messages.success(request, "Roster Imported Successfully")
        elif success_count == 0:
            logging.info("Failed to Import all Rosters")
            messages.error(request, "Failed to Import all Rosters")
    except Exception as e:
        logging.error(f"Failed to process roster file: {e}")
        messages.error(request, "Failed to process roster file")


def count_non_empty_rows(worksheet):
    return len(
        [row for row in worksheet if not all([cell.value is None for cell in row])]
    )


def process_roster_row(request, row, index):
    try:
        employee = Employee.objects.get(user__employee_id=row[1])
        logging.info(f"{employee.user.name} exists")

        # Retrieve related objects and check if any are missing
        process = get_object_or_log_error(Process, row[5], "Process", index, request)
        site = get_object_or_log_error(Site, row[0], "Site", index, request)
        work_role = get_object_or_log_error(
            WorkRole, row[7], "WorkRole", index, request
        )
        lob = get_object_or_log_error(LOB, row[6], "LOB", index, request)
        # Check if all related objects are found
        if not all([process, site, work_role, lob]):
            logging.error(f"Missing related objects for row index {index}")
            return False
        # Get shift name
        shift_name = get_shift_name(row[15], row[17])
        if not shift_name:
            message = f"Invalid shift start time type for row index {index}"
            logging.error(message)
            messages.error(request, message)
            return False

        # Get shift legend
        shift_legend = get_shift_legend(shift_name)
        if not shift_legend:
            message = f"ShiftLegend {shift_name} does not exist for row index {index}"
            logging.error(message)
            messages.error(request, message)
            return False

        supervisor_1 = getSupervisor(row[8], row[9])
        supervisor_2 = getSupervisor(row[10], row[11])

        roster_data = {
            "employee": employee,
            "start_date": row[14],
            "start_time": row[15] if shift_legend.shift_count == 1 else None,
            "end_date": row[16],
            "end_time": row[17] if shift_legend.shift_count == 1 else None,
            "shift_legend": shift_legend,
            "gender": row[12].strip(),
            "process": process,
            "site": site,
            "work_role": work_role,
            "lob": lob,
            "pick_drop_location": str(row[13]).strip(),
            "supervisor_1": supervisor_1,
            "supervisor_2": supervisor_2,
        }
        return roster_creation(roster_data, request, index)
    except Employee.DoesNotExist:
        message = f"Employee does not exist for row index {index}"
        logging.error(message)
        messages.error(request, message)
        return False


def get_object_or_none(model, name):
    try:
        return model.objects.get(name=name.lower().strip())
    except model.DoesNotExist:
        return None


def get_shift_name(start_time, end_time):
    if isinstance(start_time, time):
        return convertTimeRange(start_time, end_time)
    elif isinstance(start_time, str):
        return start_time
    return None


def get_shift_legend(shift_name):
    try:
        return ShiftLegend.objects.get(shift_name=shift_name)
    except ShiftLegend.DoesNotExist:
        return None


def log_failed_rows(failed_rows):
    if failed_rows:
        logging.info(f"Failed rows: {', '.join(map(str, failed_rows))}")
