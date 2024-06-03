from django.shortcuts import render, redirect, get_object_or_404

from rms.decorators import check_user_able_to_see_page
from .forms import *
from django.contrib.auth.forms import UserCreationForm
from .models import *
from django.contrib import messages
from openpyxl import load_workbook
import logging
from django.utils.html import format_html
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.urls import reverse
from django.template.loader import render_to_string
from django.http import JsonResponse
from .utils import *
from rms.page_info_collection import PageInfoCollection
from django.db.models import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required
from rms.constants import GroupEnum
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.forms import PasswordChangeForm
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.models import Group
from django.http import HttpResponseForbidden

logger = logging.getLogger(__name__)


class CustomPasswordChangeView(PasswordChangeView):
    # print("cpcv")
    template_name = (
        "registration/password_change.html"  # Change this to your desired template
    )
    form_class = PasswordChangeForm
    success_url = reverse_lazy(
        "home"
    )  # Change 'home' to the name of your home page URL

    def form_valid(self, form):
        # print("form valid")
        response = super().form_valid(form)
        messages.success(
            self.request, "Your password has been successfully changed."
        )  # Add success message
        # print(self.request.user.id)
        try:
            employee = Employee.objects.get(user=self.request.user.id)
            employee.is_previously_logged_in = True
            employee.save()
        except Exception as e:
            pass
            # print(f"Exception : {e}")
        return response

    def form_invalid(self, form):
        # print("form invalid")
        response = super().form_invalid(form)
        messages.error(
            self.request, "There was an error changing your password. Please try again."
        )  # Add error message
        return response


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewSettings(request):
    templateName = "accounts/view_settings.html"

    context = {
        "breadCrumb": {},
        "details": {
            "name": "delete",
            "type": "site",
        },
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewUsers(request):
    templateName = "accounts/view_users.html"

    context = {
        "breadCrumb": {},
        "details": {
            "name": "delete",
            "type": "site",
        },
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm, GroupEnum.supervisor, GroupEnum.employee)
def home(request):
    templateName = "home/view.html"
    skills = Skill.objects.all()
    if request.user.is_Employee():
        employee = Employee.objects.get(user=request.user.id)
        if employee.is_previously_logged_in == False:
            return redirect("password_change")
    context = {
        "skills": skills,
        "reportingOneUrl": PageInfoCollection.REPORTINGONE_JSON.urlName,
    }
    return render(request, templateName, context)


################################################################
#   Process
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewProcess(request):
    templateName = "process/view.html"
    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.PROCESS_VIEW]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.PROCESS_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "process",
        },
        "ajaxUrl": PageInfoCollection.PROCESS_JSON.urlName,
        "createUrl": PageInfoCollection.PROCESS_CREATE.urlName,
        "settingsActive": "active open",
        "processActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createProcess(request):
    templateName = "process/create.html"

    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.PROCESS_VIEW,
        PageInfoCollection.PROCESS_CREATE,
    ]
    if request.method == "POST":
        form = CreateProcess(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            name = data["name"].lower()
            try:
                Process.objects.get(name=name)
                error = f"Process with this name: {name.title()} already exist"
                messages.error(request, error)
                logger.error(error)
            except Process.DoesNotExist:
                try:
                    process = Process(
                        name=name,
                        created_by=CustomUser.objects.get(id=request.user.id),
                    )
                    process.save()
                    messages.success(request, "Process Created successfully.")
                    return redirect(PageInfoCollection.PROCESS_VIEW.urlName)
                except Exception as e:
                    messages.error(request, "Exception:{exception}".format(exception=e))
        else:
            messages.error(request, "Form is not valid")
    else:
        form = CreateProcess()
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.PROCESS_CREATE.pageName,
        "form": form,
        "formUrl": PageInfoCollection.PROCESS_CREATE.urlName,
        "settingsActive": "active open",
        "processActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editProcess(request, id):
    template_name = "process/edit.html"

    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.PROCESS_VIEW,
        PageInfoCollection.PROCESS_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    process = get_object_or_404(Process, id=id)

    if request.method == "POST":
        form = CreateProcess(request.POST, instance=process)
        if form.is_valid():
            data = form.cleaned_data
            name = data["name"].lower()

            try:
                # Check if a process with the same name already exists (excluding the current one)
                existing_process = Process.objects.exclude(id=process.id).get(name=name)
                error = f"Process with this name: {name.title()} already exists"
                messages.error(request, error)
            except Process.DoesNotExist:
                # Update the existing process with the new data
                process.name = name
                process.updated_by = CustomUser.objects.get(id=request.user.id)
                process.save()
                messages.success(request, "Process updated successfully.")
                return redirect("viewProcess")
        else:
            messages.error(request, "Form is not valid")
    else:
        # If it's a GET request, populate the form with the existing process data
        form = CreateProcess(instance=process)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.PROCESS_EDIT.pageName,
        "formUrl": PageInfoCollection.PROCESS_EDIT.urlName,
        "settingsActive": "active open",
        "processActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteProcess(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            process = Process.objects.get(id=id)
            process.delete()
            success = True
        except Process.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Process |id:{id}| Exception: Process does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Process |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewProcessJson(BaseDatatableView):
    model = Process

    # Define the columns you want to display
    columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            return Process.objects.all().order_by("name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = [
                "name",
                "created_by__name",
                "created_at",
                "updated_by__name",
                "updated_at",
            ]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.title(),
                "created_by": "" if item.created_by is None else item.created_by.name,
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_at": item.updated_at.strftime("%d-%b-%y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewProcessJson, self).render_column(row, column)

    def get_actions_html(self, item):
        # Implement your action links here
        edit_url = reverse("editRoster", args=[item.id])
        edit_link = f'<a href="{edit_url}" class="btn btn-success">Edit</a>'
        return f"{edit_link}"

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.PROCESS_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.PROCESS_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


################################################################
#   Site                                                       #
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewSite(request):
    templateName = "site/view.html"

    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.SITE_VIEW]

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SITE_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "site",
        },
        "ajaxUrl": PageInfoCollection.SITE_JSON.urlName,
        "createUrl": PageInfoCollection.SITE_CREATE.urlName,
        "settingsActive": "active open",
        "siteActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createSite(request):
    templateName = "site/create.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.SITE_VIEW,
        PageInfoCollection.SITE_CREATE,
    ]
    if request.method == "POST":
        result = createSingleNameModel(request, Site, CreateSite)
        if result == True:
            return redirect(PageInfoCollection.SITE_VIEW.urlName)
        else:
            form = CreateSite(request.POST)
    else:
        form = CreateSite()
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SITE_CREATE.pageName,
        "form": form,
        "formUrl": PageInfoCollection.SITE_CREATE.urlName,
        "settingsActive": "active open",
        "siteActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editSite(request, id):
    template_name = "site/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.SITE_VIEW,
        PageInfoCollection.SITE_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    site = get_object_or_404(Site, id=id)

    if request.method == "POST":
        form = CreateSite(request.POST, instance=site)
        if form.is_valid():
            data = form.cleaned_data
            name = data["name"].lower()

            try:
                # Check if a process with the same name already exists (excluding the current one)
                existing_site = Site.objects.exclude(id=site.id).get(name=name)
                error = f"Site with this name: {name.title()} already exists"
                messages.error(request, error)
            except Site.DoesNotExist:
                # Update the existing process with the new data
                site.name = name
                site.updated_by = CustomUser.objects.get(id=request.user.id)
                site.save()
                messages.success(request, "Site updated successfully.")
                return redirect(PageInfoCollection.SITE_VIEW.urlName)
        else:
            messages.error(request, "Form is not valid")
    else:
        # If it's a GET request, populate the form with the existing process data
        form = CreateProcess(instance=site)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SITE_EDIT.pageName,
        "formUrl": PageInfoCollection.SITE_EDIT.urlName,
        "settingsActive": "active open",
        "siteActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteSite(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            site = Site.objects.get(id=id)
            site.delete()
            success = True
        except Process.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Site |id:{id}| Exception: Process does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Site |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewSiteJson(BaseDatatableView):
    model = Site

    # Define the columns you want to display
    columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            return Site.objects.all().order_by("name")
        return Site.objects.none()

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = [
                "name",
                "created_by__name",
                "created_at",
                "updated_by__name",
                "updated_at",
            ]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.title(),
                "created_by": "" if item.created_by is None else item.created_by.name,
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_at": (
                    ""
                    if item.updated_at is None
                    else item.updated_at.strftime("%d-%b-%y")
                ),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewSiteJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.SITE_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.SITE_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


################################################################
#   LOB
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewLOB(request):
    templateName = "lob/view.html"

    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.LOB_VIEW]

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.LOB_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "LOB",
        },
        "ajaxUrl": PageInfoCollection.LOB_JSON.urlName,
        "createUrl": PageInfoCollection.LOB_CREATE.urlName,
        "settingsActive": "active open",
        "lobActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createLOB(request):
    templateName = "lob/create.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.LOB_VIEW,
        PageInfoCollection.LOB_CREATE,
    ]
    if request.method == "POST":
        result = createSingleNameModel(request, LOB, CreateLOB)
        if result == True:
            return redirect(PageInfoCollection.LOB_VIEW.urlName)
        else:
            form = CreateLOB(request.POST)
    else:
        form = CreateLOB()
    context = {
        "form": form,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.LOB_CREATE.pageName,
        "formUrl": PageInfoCollection.LOB_CREATE.urlName,
        "settingsActive": "active open",
        "lobActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editLOB(request, id):
    template_name = "lob/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.LOB_VIEW,
        PageInfoCollection.LOB_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    lob = get_object_or_404(LOB, id=id)

    if request.method == "POST":
        form = CreateLOB(request.POST, instance=lob)
        if form.is_valid():
            data = form.cleaned_data
            name = data["name"].lower()

            try:
                # Check if a process with the same name already exists (excluding the current one)
                existing_LOB = LOB.objects.exclude(id=lob.id).get(name=name)
                error = f"LOB with this name: {name.title()} already exists"
                messages.error(request, error)
            except LOB.DoesNotExist:
                # Update the existing process with the new data
                lob.name = name
                lob.updated_by = CustomUser.objects.get(id=request.user.id)
                lob.save()
                messages.success(request, "LOB updated successfully.")
                return redirect(PageInfoCollection.LOB_VIEW.urlName)
        else:
            messages.error(request, "Form is not valid")
    else:
        # If it's a GET request, populate the form with the existing process data
        form = CreateProcess(instance=lob)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.LOB_EDIT.pageName,
        "formUrl": PageInfoCollection.LOB_EDIT.urlName,
        "settingsActive": "active open",
        "lobActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteLOB(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            lob = LOB.objects.get(id=id)
            lob.delete()
            success = True
        except LOB.DoesNotExist:
            logging.error(
                f"|Failed| Delete A LOB |id:{id}| Exception: LOB does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A LOB |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewLOBJson(BaseDatatableView):
    model = LOB

    # Define the columns you want to display
    columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            return LOB.objects.all().order_by("name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = [
                "name",
                "created_by__name",
                "created_at",
                "updated_by__name",
                "updated_at",
            ]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.upper(),
                "created_by": "" if item.created_by is None else item.created_by.name,
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "updated_at": item.updated_at.strftime("%d-%b-%y"),
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewLOBJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.LOB_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.LOB_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


################################################################
#   WorkRole
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewWorkRole(request):
    templateName = "workRole/view.html"

    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.WORKROLE_VIEW]

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.WORKROLE_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Work Role",
        },
        "ajaxUrl": PageInfoCollection.WORKROLE_JSON.urlName,
        "createUrl": PageInfoCollection.WORKROLE_CREATE.urlName,
        "settingsActive": "active open",
        "workRoleActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createWorkRole(request):
    templateName = "workRole/create.html"

    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.WORKROLE_VIEW,
        PageInfoCollection.WORKROLE_CREATE,
    ]

    if request.method == "POST":
        result = createSingleNameModel(request, WorkRole, CreateWorkRole)
        if result == True:
            return redirect(PageInfoCollection.WORKROLE_VIEW.urlName)
        else:
            form = CreateWorkRole(request.POST)
    else:
        form = CreateWorkRole()
    context = {
        "form": form,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.WORKROLE_CREATE.pageName,
        "formUrl": PageInfoCollection.WORKROLE_CREATE.urlName,
        "settingsActive": "active open",
        "workRoleActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editWorkRole(request, id):
    template_name = "workRole/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.WORKROLE_VIEW,
        PageInfoCollection.WORKROLE_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    workRole = get_object_or_404(WorkRole, id=id)

    if request.method == "POST":
        form = CreateWorkRole(request.POST, instance=workRole)
        if form.is_valid():
            data = form.cleaned_data
            name = data["name"].lower()

            try:
                # Check if a process with the same name already exists (excluding the current one)
                existing_WorkRole = WorkRole.objects.exclude(id=workRole.id).get(
                    name=name
                )
                error = f"WorkRole with this name: {name.title()} already exists"
                messages.error(request, error)
            except LOB.DoesNotExist:
                # Update the existing process with the new data
                workRole.name = name
                workRole.updated_by = CustomUser.objects.get(id=request.user.id)
                workRole.save()
                messages.success(request, "WorkRole updated successfully.")
                return redirect(PageInfoCollection.WORKROLE_VIEW.urlName)
        else:
            messages.error(request, "Form is not valid")
    else:
        # If it's a GET request, populate the form with the existing process data
        form = CreateProcess(instance=workRole)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.WORKROLE_EDIT.pageName,
        "formUrl": PageInfoCollection.WORKROLE_EDIT.urlName,
        "settingsActive": "active open",
        "workRoleActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteWorkRole(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            workRole = WorkRole.objects.get(id=id)
            workRole.delete()
            success = True
        except LOB.DoesNotExist:
            logging.error(
                f"|Failed| Delete A WorkRole |id:{id}| Exception: LOB does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A WorkRole |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewWorkRoleJson(BaseDatatableView):
    model = WorkRole

    # Define the columns you want to display
    columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            return WorkRole.objects.all().order_by("name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = [
                "name",
                "created_by__name",
                "created_at",
                "updated_by__name",
                "updated_at",
            ]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.upper(),
                "created_by": "" if item.created_by is None else item.created_by.name,
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "updated_at": item.updated_at.strftime("%d-%b-%y"),
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewWorkRoleJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.WORKROLE_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.WORKROLE_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


################################################################
#   Skill
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewSkill(request):
    templateName = "skill/view.html"

    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.SKILL_VIEW]

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SKILL_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Skill",
        },
        "ajaxUrl": PageInfoCollection.SKILL_JSON.urlName,
        "createUrl": PageInfoCollection.SKILL_CREATE.urlName,
        "settingsActive": "active open",
        "skillActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createSkill(request):
    templateName = "skill/create.html"

    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.SKILL_VIEW,
        PageInfoCollection.SKILL_CREATE,
    ]

    if request.method == "POST":
        result = createSingleNameModel(request, Skill, CreateSkill)
        if result == True:
            return redirect(PageInfoCollection.SKILL_VIEW.urlName)
        else:
            form = CreateSkill(request.POST)
    else:
        form = CreateSkill()
    context = {
        "form": form,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SKILL_CREATE.pageName,
        "formUrl": PageInfoCollection.SKILL_CREATE.urlName,
        "settingsActive": "active open",
        "skillActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editSkill(request, id):
    template_name = "skill/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.SKILL_VIEW,
        PageInfoCollection.SKILL_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    skill = get_object_or_404(Skill, id=id)

    if request.method == "POST":
        form = CreateWorkRole(request.POST, instance=skill)
        if form.is_valid():
            data = form.cleaned_data
            name = data["name"].lower()

            try:
                # Check if a process with the same name already exists (excluding the current one)
                existing_skill = Skill.objects.exclude(id=skill.id).get(name=name)
                error = f"Skill with this name: {name.title()} already exists"
                messages.error(request, error)
            except Skill.DoesNotExist:
                # Update the existing process with the new data
                skill.name = name
                skill.updated_by = CustomUser.objects.get(id=request.user.id)
                skill.save()
                messages.success(request, "Skill updated successfully.")
                return redirect(PageInfoCollection.SKILL_VIEW.urlName)
        else:
            messages.error(request, "Form is not valid")
    else:
        # If it's a GET request, populate the form with the existing process data
        form = CreateProcess(instance=skill)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SKILL_EDIT.pageName,
        "formUrl": PageInfoCollection.SKILL_EDIT.urlName,
        "settingsActive": "active open",
        "skillActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteSkill(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            skill = Skill.objects.get(id=id)
            skill.delete()
            success = True
        except LOB.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Skill |id:{id}| Exception: LOB does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Skill |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewSkillJson(BaseDatatableView):
    model = Skill

    # Define the columns you want to display
    columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            return Skill.objects.all().order_by("name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = [
                "name",
                "created_by__name",
                "created_at",
                "updated_by__name",
                "updated_at",
            ]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.upper(),
                "created_by": "" if item.created_by is None else item.created_by.name,
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_by": "" if item.updated_by is None else item.updated_by.name,
                "updated_at": item.updated_at.strftime("%d-%b-%y"),
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewSkillJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.SKILL_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.SKILL_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


################################################################
#   Supervisor
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewSupervisor(request):
    templateName = "supervisor/view.html"
    breadCrumbList = [PageInfoCollection.USERS, PageInfoCollection.SUPERVISOR_VIEW]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SUPERVISOR_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "supervisor",
        },
        "ajaxUrl": PageInfoCollection.SUPERVISOR_JSON.urlName,
        "createUrl": PageInfoCollection.SUPERVISOR_CREATE.urlName,
        "bulkUrl": PageInfoCollection.SUPERVISOR_BULK.urlName,
        "supervisorActive": "active",
        "usersActive": "active open",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createSupervisor(request):
    template_name = "supervisor/create.html"
    breadCrumbList = [
        PageInfoCollection.USERS,
        PageInfoCollection.SUPERVISOR_VIEW,
        PageInfoCollection.SUPERVISOR_CREATE,
    ]
    if request.method == "POST":
        form = CreateEditSupervisorForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            # print(data)
            result = supervisorCreation(data, request, False)
            if result == True:
                return redirect(PageInfoCollection.SUPERVISOR_VIEW.urlName)
        else:
            messages.error(request, "Could not create Supervisor")
    else:
        form = CreateEditSupervisorForm()
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SUPERVISOR_CREATE.pageName,
        "form": form,
        "formUrl": PageInfoCollection.SUPERVISOR_CREATE.urlName,
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editSupervisor(request, id):
    template_name = "supervisor/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.SUPERVISOR_VIEW,
        PageInfoCollection.SUPERVISOR_EDIT,
    ]
    # Retrieve the existing supervisor based on the provided id
    supervisor = get_object_or_404(Supervisor, id=id)

    if request.method == "POST":
        form = CreateEditSupervisorForm(request.POST, instance=supervisor)

        if form.is_valid():
            data = form.cleaned_data
            email = data["email"].lower()

            try:
                # Check if a supervisor with the same name already exists (excluding the current one)
                existing_supervisor = Supervisor.objects.exclude(id=supervisor.id).get(
                    user__email=email
                )
                error = f"Supervisor with this email: {email} already exists"
                messages.error(request, error)
            except Supervisor.DoesNotExist:
                # Update the existing supervisor with the new data
                # print(data)
                customUser = CustomUser.objects.get(id=supervisor.user.id)
                process = Process.objects.get(name=data["process"])
                site = Site.objects.get(name=data["site"])
                workRole = WorkRole.objects.get(name=data["work_role"])

                customUser.email = data["email"]
                customUser.name = data["name"]
                customUser.system_id = data["system_id"]
                customUser.employee_id = data["employee_id"]
                customUser.save()

                supervisor.process = process
                supervisor.site = site
                supervisor.workRole = workRole
                supervisor.gender = data["gender"]
                supervisor.updated_by = request.user
                supervisor.save()
                messages.success(request, "Supervisor updated successfully.")
                return redirect(PageInfoCollection.SUPERVISOR_VIEW.urlName)
        else:
            messages.error(request, "Form is not valid")
    else:
        form = CreateEditSupervisorForm(instance=supervisor)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SUPERVISOR_EDIT.pageName,
        "formUrl": PageInfoCollection.SUPERVISOR_EDIT.urlName,
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteSupervisor(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            supervisor = Supervisor.objects.get(id=id)
            supervisor.delete()
            success = True
        except Supervisor.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Supervisor |id:{id}| Exception: LOB does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Supervisor |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class ViewSupervisorJson(BaseDatatableView):
    model = Supervisor

    # Define the columns you want to display
    columns = [
        "user__name",
        "process__name",
        "gender",
        "site__name",
        "work_role__name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "user__name",
        "process__name",
        "gender",
        "site__name",
        "work_role__name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            return Supervisor.objects.all().order_by("user__name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        search = self.request.POST.get("search[value]", None)
        if search:
            # Filter by employee's name
            qs = qs.filter(employee__user__name__icontains=search)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "user__name": item.user.name.title(),
                "process__name": item.process.name.title(),
                "gender": "Male" if item.gender is "M" else "Female",
                "site__name": item.site.name.capitalize() if item.site else "",
                "work_role__name": (
                    item.work_role.name.title() if item.work_role else ""
                ),
                "created_by": (
                    "" if item.created_by is None else item.created_by.name.title()
                ),
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_by": (
                    "" if item.updated_by is None else item.updated_by.name.title()
                ),
                "updated_at": item.updated_at.strftime("%d-%b-%y"),
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(ViewSupervisorJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.SUPERVISOR_EDIT.urlName, args=[item.id])
        reject_url = reverse(
            PageInfoCollection.SUPERVISOR_DELETE.urlName, args=[item.id]
        )

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def bulkAddSupervisors(request):
    template_name = "supervisor/bulk.html"
    breadCrumbList = [
        PageInfoCollection.USERS,
        PageInfoCollection.SUPERVISOR_VIEW,
        PageInfoCollection.SUPERVISOR_BULK,
    ]
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        ws = wb["Sheet1"]
        count = (
            len([row for row in ws if not all([cell.value == None for cell in row])])
            - 1
        )
        successCount = 0
        failedList = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                # print("row:{name}".format(name=row[0]))
                data = {
                    "name": row[0].lower(),
                    "email": row[1].lower(),
                    #'email':"{preHeader}@genex.com".format(row[0].replace(" ", "")),
                    "employee_id": row[2],
                    "system_id": row[3],
                    "gender": row[4],
                    "process": Process.objects.get(name=row[5].lower()),
                    "site": Site.objects.get(name=row[6].lower()),
                    "work_role": WorkRole.objects.get(name=row[7].lower()),
                    "password1": 1234,
                }
                result = supervisorCreation(data, request, True)
                if result == True:
                    successCount += 1
                else:
                    failedList.append(row[0])

        if count == successCount:
            messages.success(request, "Supervisors Created successfully.")
        else:
            messages.error(request, f"Failed : {','.join(failedList)}")
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.SUPERVISOR_BULK.pageName,
        "formUrl": PageInfoCollection.SUPERVISOR_BULK.urlName,
    }
    return render(request, template_name, context)


################################################################
#   Employee
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewEmployee(request):
    templateName = "employee/view.html"
    breadCrumbList = [PageInfoCollection.USERS, PageInfoCollection.EMPLOYEE_VIEW]
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.EMPLOYEE_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "employee",
        },
        "ajaxUrl": PageInfoCollection.EMPLOYEE_JSON.urlName,
        "createUrl": PageInfoCollection.EMPLOYEE_CREATE.urlName,
        "otherInfoUrl": PageInfoCollection.EMPLOYEE_OTHERINFO.urlName,
        "bulkUrl": PageInfoCollection.EMPLOYEE_BULK.urlName,
        "usersActive": "active open",
        "employeeActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createEmployee(request):
    template_name = "employee/create.html"
    breadCrumbList = [
        PageInfoCollection.USERS,
        PageInfoCollection.EMPLOYEE_VIEW,
        PageInfoCollection.EMPLOYEE_CREATE,
    ]
    if request.method == "POST":
        form = CreateEditEmployeeForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            result = employeeCreation(data, request, False, data["userType"])
            if result == True:
                return redirect(PageInfoCollection.EMPLOYEE_VIEW.urlName)
        else:
            # Access form errors
            errors = form.errors
            logging.error(f"|Failed| Create An Employee |id:{id}| Exception: {errors}.")
            messages.error(request, errors)
    else:
        form = CreateEditEmployeeForm()
    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.EMPLOYEE_CREATE.pageName,
        "form": form,
        "formUrl": PageInfoCollection.EMPLOYEE_CREATE.urlName,
        "usersActive": "active open",
        "employeeActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editEmployee(request, id):
    template_name = "employee/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.EMPLOYEE_VIEW,
        PageInfoCollection.EMPLOYEE_EDIT,
    ]
    # Retrieve the existing EMPLOYEE based on the provided id
    employee = get_object_or_404(Employee, id=id)

    if request.method == "POST":
        form = CreateEditEmployeeForm(request.POST, instance=employee)

        if form.is_valid():
            data = form.cleaned_data
            email = data["email"].lower()

            try:
                # Check if a Employee with the same name already exists (excluding the current one)
                existing_employee = Employee.objects.exclude(id=employee.id).get(
                    user__email=email
                )
                error = f"Employee with this email: {email} already exists"
                messages.error(request, error)
            except Employee.DoesNotExist:
                # Update the existing Employee with the new data
                # print(data)
                try:
                    customUser = CustomUser.objects.get(id=employee.user.id)
                    process = (
                        None
                        if data["process"] is None
                        else Process.objects.get(name=data["process"])
                    )
                    site = (
                        None
                        if data["site"] is None
                        else Site.objects.get(name=data["site"])
                    )
                    workRole = (
                        None
                        if data["work_role"] is None
                        else WorkRole.objects.get(name=data["work_role"])
                    )
                    lob = (
                        None
                        if data["lob"] is None
                        else LOB.objects.get(name=data["lob"])
                    )

                    customUser.email = data["email"]
                    customUser.name = data["name"]
                    customUser.system_id = data["system_id"]
                    customUser.employee_id = data["employee_id"]
                    customUser.is_active = data["is_active"]
                    customUser.save()

                    employee.process = process
                    employee.site = site
                    employee.workRole = workRole
                    employee.gender = data["gender"]
                    employee.lob = lob
                    employee.pick_drop_location = data["pick_drop_location"]
                    employee.supervisor_1 = data["supervisor_1"]
                    employee.supervisor_2 = data["supervisor_2"]
                    employee.updated_by = request.user
                    employee.save()
                    logging.info(f"|Success| Edit A Employee |id:{id}|")
                    messages.success(request, "Employee updated successfully.")
                except Exception as e:
                    logging.error(f"|Failed| Edit A Employee |id:{id}| Exception:{e}")
                    messages.error(request, "Failed to edit employee.")

                return redirect(PageInfoCollection.EMPLOYEE_VIEW.urlName)
        else:
            messages.error(request, "Form is not valid")
    else:
        form = CreateEditEmployeeForm(instance=employee)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.EMPLOYEE_EDIT.pageName,
        "formUrl": PageInfoCollection.EMPLOYEE_EDIT.urlName,
        "usersActive": "active open",
        "employeeActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteEmployee(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            employee = Employee.objects.get(id=id)
            employee.delete()
            success = True
        except Employee.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Employee |id:{id}| Exception: Employee does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Employee |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class ViewEmployeeJson(BaseDatatableView):
    model = Employee

    # Define the columns you want to display
    columns = [
        "user__name",
        "user__email",
        "user__employee_id",
        "user__system_id",
        "user__is_active",
        "avaya_id",
        "vdi",
        "doj",
        "process__name",
        "gender",
        "site__name",
        "work_role__name",
        "lob__name",
        "pick_drop_location",
        "supervisor_1__user__name",
        "supervisor_2__user__name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "user__name",
        "user__email",
        "user__employee_id",
        "user__system_id",
        "user__is_active",
        "avaya_id",
        "vdi",
        "doj",
        "process__name",
        "gender",
        "site__name",
        "work_role__name",
        "lob__name",
        "pick_drop_location",
        "supervisor_1__user__name",
        "supervisor_2__user__name",
        "created_by",
        "created_at",
        "updated_by",
        "updated_at",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request

        if (
            self.request.user.is_WFM()
            or self.request.user.is_Employee()
            or self.request.user.is_Supervisor()
        ):
            userType = self.request.GET.get("userType")
            supervisor_group = Group.objects.get(name="Supervisor")
            if userType == "Supervisors":
                return Employee.objects.filter(
                    Q(work_role__name="Supervisor") | Q(user__groups=supervisor_group)
                ).order_by("user__name")
            elif userType == "Agents":
                return (
                    Employee.objects.all()
                    .exclude(
                        Q(work_role__name="Supervisor")
                        | Q(user__groups=supervisor_group)
                    )
                    .order_by("user__name")
                )
            else:
                return Employee.objects.all().order_by("user__name")
        else:
            Employee.objects.none()

    def filter_queryset(self, qs):
        # Handle search parameter from DataTables
        search_value = self.request.GET.get("search[value]", None)

        if search_value:
            # Define the fields you want to search on
            search_fields = [
                "user__name",
                "user__email",
                "user__employee_id",
                "user__system_id",
                "user__is_active",
                "avaya_id",
                "vdi",
                "doj",
                "process__name",
                "gender",
                "site__name",
                "work_role__name",
                "lob__name",
                "pick_drop_location",
                "supervisor_1__user__name",
                "supervisor_2__user__name",
                "created_by__name",
                "created_at",
                "updated_by__name",
                "updated_at",
            ]

            # Create a Q object to dynamically construct the filter conditions
            search_filter = Q()
            for field in search_fields:
                search_filter |= Q(**{f"{field}__icontains": search_value})

            # Apply the search filter to the queryset
            qs = qs.filter(search_filter)

        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "user__name": "" if item.user.name is None else item.user.name.title(),
                "user__email": item.user.email,
                "user__employee_id": item.user.employee_id,
                "user__system_id": item.user.system_id,
                "user__is_active": item.user.is_active,
                "avaya_id": "" if item.avaya_id is None else item.avaya_id,
                "vdi": "" if item.vdi is None else item.vdi,
                "doj": "" if item.doj is None else item.doj.strftime("%d-%b-%y"),
                "process__name": (
                    "" if item.process is None else item.process.name.title()
                ),
                "gender": "Male" if item.gender is "M" else "Female",
                "site__name": item.site.name.capitalize() if item.site else "",
                "work_role__name": (
                    item.work_role.name.title() if item.work_role else ""
                ),
                "lob__name": item.lob.name.title() if item.lob else "",
                "pick_drop_location": (
                    item.pick_drop_location.title() if item.pick_drop_location else ""
                ),
                "supervisor_1__user__name": (
                    item.supervisor_1.user.name.title() if item.supervisor_1 else ""
                ),
                "supervisor_2__user__name": (
                    item.supervisor_2.user.name.title() if item.supervisor_2 else ""
                ),
                "created_by": (
                    "" if item.created_by is None else item.created_by.name.title()
                ),
                "created_at": item.created_at.strftime("%d-%b-%y"),
                "updated_by": (
                    "" if item.updated_by is None else item.updated_by.name.title()
                ),
                "updated_at": item.updated_at.strftime("%d-%b-%y"),
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(ViewEmployeeJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.EMPLOYEE_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.EMPLOYEE_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def bulkAddEmployees(request):
    template_name = "employee/bulk.html"
    breadCrumbList = [
        PageInfoCollection.USERS,
        PageInfoCollection.EMPLOYEE_VIEW,
        PageInfoCollection.EMPLOYEE_BULK,
    ]

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if "supervisor" in request.POST:
            selectedGroup = "Supervisor"
        else:
            selectedGroup = "Employee"

        logger.info(f"Selected Group: {selectedGroup}")
        if excel_file:
            wb = load_workbook(excel_file)
            ws = wb["Sheet1"]
            count = (
                len(
                    [row for row in ws if not all([cell.value == None for cell in row])]
                )
                - 1
            )
            # print(count)
            successCount = 0
            failedList = []
            logger.info(f"Trying to insert data")
            for index, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                logger.info(f"--------------------------------")
                if row[0] is not None:
                    logger.info(f"Index:{index}")
                    flagEmptyFields = False
                    customUserData = {
                        "name": row[0],
                        # "email": row[1],
                        "email": (
                            f"{row[0].lower().replace(' ','')}{row[2]}@zohomail.com"
                            if row[1] is None
                            else row[1]
                        ),
                        "employee_id": row[2],
                        "system_id": row[3],
                        "gender": row[4],
                    }

                    for key, value in customUserData.items():
                        if value is None:
                            logger.error(f"{key} is None.")
                            messages.error(
                                request, f"Failed : Name: {row[0]}. {key} is None."
                            )
                            flagEmptyFields = True

                    if not flagEmptyFields:
                        logger.info(f"No Empty Fields in customUserData")
                        logger.info(f"Employee Name: {row[0]}")
                        try:
                            pick_drop_location = row[9]

                            supervisor1_email = row[11]
                            supervisor2_email = row[13]
                            logger.info(f"supervisor1 email: {supervisor1_email}")
                            supervisor1 = None
                            if supervisor1_email is not None:
                                try:
                                    supervisor1 = Employee.objects.get(
                                        user__email=supervisor1_email.lower().strip()
                                    )
                                except Employee.DoesNotExist:
                                    logger.error(
                                        f"Supervisor [{supervisor1_email}] does not exist"
                                    )

                            logger.info(f"supervisor1: {supervisor1}")
                            # supervisor2 = (
                            #     Employee.objects.get(
                            #         user__email=supervisor2_email.lower()
                            #     )
                            #     if supervisor2_email
                            #     else None
                            # )
                            supervisor2 = None

                            password = 123456
                            tempLob = str(row[8])
                            employeeData = {
                                "process": row[5],
                                "site": row[6],
                                "work_role": row[7],
                                "lob": row[8],
                                "pick_drop_location": pick_drop_location,
                                "supervisor_1": supervisor1,
                                "supervisor_2": supervisor2,
                                "password1": password,
                            }

                            if not flagEmptyFields:
                                data = {**customUserData, **employeeData}
                                result = employeeCreation(
                                    data, request, True, selectedGroup
                                )

                                if result:
                                    successCount += 1
                                else:
                                    failedList.append(row[0])

                        # except ObjectDoesNotExist as e:
                        #     logging.error(f"|ObjectDoesNotExist| {e}")
                        #     flagEmptyFields = True
                        except Exception as e:
                            logging.error(
                                f"|Failed| [Index:{index}] Create A Employee in bulk |id:{row[2]}| Exception:{e}"
                            )
                    else:
                        logger.error(
                            f"|Failed| [Index:{index}] Empty Fields in customUserData"
                        )
                logger.info(f"--------------------------------")
            if count == successCount:
                logger.info(f"|Success| All Employees Created successfully")
                messages.success(request, "Employees Created successfully.")
            else:
                logger.error(f"|Failed| {','.join(failedList)}")
                messages.error(request, f"Failed : {','.join(failedList)}")

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.EMPLOYEE_BULK.pageName,
        "formUrl": PageInfoCollection.EMPLOYEE_BULK.urlName,
        "usersActive": "active open",
        "employeeActive": "active",
    }

    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def bulkAddEmployeesOtherInfo(request):
    template_name = "employee/otherInfo.html"
    breadCrumbList = [
        PageInfoCollection.USERS,
        PageInfoCollection.EMPLOYEE_VIEW,
        PageInfoCollection.EMPLOYEE_OTHERINFO,
    ]

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if excel_file:
            wb = load_workbook(excel_file)
            ws = wb["Sheet1"]
            count = (
                len(
                    [row for row in ws if not all([cell.value == None for cell in row])]
                )
                - 1
            )
            # print(count)
            successCount = 0
            failedList = []
            for index, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if row[0] is not None:
                    try:
                        employee = Employee.objects.get(user__employee_id=row[0])
                        # employee.vdi = row[1]
                        # employee.avaya_id = row[2]
                        # employee.doj = row[4]
                        # ---
                        employee.vdi = row[3]
                        employee.avaya_id = row[2]
                        # employee.doj = row[4]
                        employee.save()
                        successCount += 1
                        logger.info(
                            f"{employee.user.name} other info added successfully"
                        )
                    except Employee.DoesNotExist:
                        logger.error(
                            f"|Failed| [Index:{index}] Exception: [Employee ID:{row[0]}] [Employee:{row[1]}] [Avaya ID:{row[2]}] [VDI ID:{row[3]}] does not exist"
                        )
                        # Handle the case where the employee is not found
                    except Exception as e:
                        logger.error(f"|Failed| [Index:{index}] Exception: {e}")
            # print(f"succes count:{successCount}")
            if count == successCount:
                messages.success(request, "Employees Created successfully.")
            else:
                messages.error(request, f"Failed : {','.join(failedList)}")

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.EMPLOYEE_OTHERINFO.pageName,
        "formUrl": PageInfoCollection.EMPLOYEE_OTHERINFO.urlName,
        "usersActive": "active open",
        "employeeActive": "active",
    }

    return render(request, template_name, context)


################################################################
#   Groups
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewGroup(request):
    templateName = "group/view.html"

    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.GROUP_VIEW]

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.GROUP_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "Group",
        },
        "ajaxUrl": PageInfoCollection.GROUP_JSON.urlName,
        "createUrl": PageInfoCollection.GROUP_CREATE.urlName,
        "settingsActive": "active open",
        "groupActive": "active",
    }
    print(request.user.groups.filter(permissions__codename="view_group"))
    if not request.user.groups.filter(permissions__codename="view_group").exists():
        return HttpResponseForbidden("You do not have permission to view this roster.")

    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createGroup(request):
    templateName = "group/create.html"

    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.GROUP_VIEW,
        PageInfoCollection.GROUP_CREATE,
    ]

    if request.method == "POST":
        form = GroupForm(request.POST)
        if form.is_valid():
            groupName = form.cleaned_data["name"]
            group = form.save()
            # Save Permissions
            group.permissions.set(form.cleaned_data["permissions"])
            group.save()
            messages.success(request, f"Group {groupName} Created Successfully")
            logging.info(f"Group {groupName} created Successfully")
            return redirect(PageInfoCollection.GROUP_VIEW.urlName)
        else:
            messages.error(
                request, "Error creating group. Please check the form for errors"
            )
            logging.error("|Failed| Error creating group: %s", form.errors)
    else:
        form = GroupForm()
    context = {
        "form": form,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.GROUP_CREATE.pageName,
        "formUrl": PageInfoCollection.GROUP_CREATE.urlName,
        "settingsActive": "active open",
        "groupActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editGroup(request, id):
    template_name = "group/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.GROUP_VIEW,
        PageInfoCollection.GROUP_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    group = get_object_or_404(Group, id=id)

    if request.method == "POST":
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            group = form.save(commit=False)
            group.permissions.set(form.cleaned_data["permissions"])
            group.save()
            messages.success(request, f"Group: {group.name} edited Successfully")
            logging.info(f"Group: {group.name} edited Successfully")
            return redirect(PageInfoCollection.GROUP_VIEW.urlName)
        else:
            messages.error(
                request, "Error updating group. Please check the form for errors"
            )
            logging.error("|Failed| Error updating group: %s", form.errors)
    else:
        # If it's a GET request, populate the form with the existing process data
        form = GroupForm(
            instance=group, initial={"permissions": group.permissions.all()}
        )

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.GROUP_EDIT.pageName,
        "formUrl": PageInfoCollection.GROUP_EDIT.urlName,
        "settingsActive": "active open",
        "groupActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteGroup(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            group = Group.objects.get(id=id)
            group.delete()
            success = True
        except Group.DoesNotExist:
            logging.error(
                f"|Failed| Delete A Group |id:{id}| Exception: Group does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A Group |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewGroupJson(BaseDatatableView):
    model = Group

    # Define the columns you want to display
    columns = [
        "name",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return Group.objects.all().order_by("name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = [
                "name",
            ]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.upper(),
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewSkillJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.GROUP_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.GROUP_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)


################################################################
#   Custom User
################################################################


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def viewUser(request):
    templateName = "user/view.html"

    breadCrumbList = [PageInfoCollection.SETTINGS, PageInfoCollection.USER_VIEW]

    context = {
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.USER_VIEW.pageName,
        "details": {
            "name": "delete",
            "type": "User",
        },
        "ajaxUrl": PageInfoCollection.USER_JSON.urlName,
        "createUrl": PageInfoCollection.USER_CREATE.urlName,
        "usersActive": "active open",
        "userActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def createUser(request):
    templateName = "user/create.html"

    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.USER_VIEW,
        PageInfoCollection.USER_CREATE,
    ]

    if request.method == "POST":
        form = CustomUserFormWithGroup(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            group = form.cleaned_data["groups"]
            if CustomUser.objects.filter(email=email).exists():
                logging.error(f"User with email '{email}' already exists")
                messages.error(
                    request, f"|Failed| User with email '{email}' already exists"
                )
            else:
                try:
                    user = form.save(commit=False)
                    user.created_by = request.user
                    userName = form.cleaned_data["name"]
                    user.set_password("123456")  # Set the password
                    user.save()
                    user.groups.set(group)
                    messages.success(request, f"User: {userName} Created Successfully")
                    logging.info(f"User: {userName} created Successfully")
                    return redirect(PageInfoCollection.USER_VIEW.urlName)
                except Exception as e:
                    messages.error(
                        request, f"An error occurred while creating the user: {e}"
                    )
                    logging.error(f"Error creating user: {e}")
        else:
            messages.error(
                request, "Error creating user. Please check the form for errors"
            )
            logging.error("|Failed| Error creating user: %s", form.errors)
    else:
        form = CustomUserFormWithGroup()
    context = {
        "form": form,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.USER_CREATE.pageName,
        "formUrl": PageInfoCollection.USER_CREATE.urlName,
        "settingsActive": "active open",
        "userActive": "active",
    }
    return render(request, templateName, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def editUser(request, id):
    template_name = "group/edit.html"
    breadCrumbList = [
        PageInfoCollection.SETTINGS,
        PageInfoCollection.USER_VIEW,
        PageInfoCollection.USER_EDIT,
    ]
    # Retrieve the existing process based on the provided process_id
    user = get_object_or_404(CustomUser, id=id)

    if request.method == "POST":
        form = CustomUserFormWithGroup(request.POST, instance=user)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f"User: {user.name} edited Successfully")
                logging.info(f"User: {user.name} edited Successfully")
                return redirect(PageInfoCollection.USER_VIEW.urlName)
            except Exception as e:
                messages.error(request, f"Error updating user: {e}")
                logger.error(f"|Failed| Error updating user: {e}")
        else:
            messages.error(
                request, "Error updating group. Please check the form for errors"
            )
            logging.error("|Failed| Error updating group: %s", form.errors)
    else:
        # If it's a GET request, populate the form with the existing process data
        form = CustomUserFormWithGroup(instance=user)

    context = {
        "form": form,
        "id": id,
        "breadCrumbList": breadCrumbList,
        "currentBreadCrumb": PageInfoCollection.USER_EDIT.pageName,
        "formUrl": PageInfoCollection.USER_EDIT.urlName,
        "settingsActive": "active open",
        "skillActive": "active",
    }
    return render(request, template_name, context)


@login_required(login_url="/login/")
@check_user_able_to_see_page(GroupEnum.wfm)
def deleteUser(request, id):
    success = False
    errorMessage = "Failed To Delete"
    successMessage = "Deleted Successfully"
    if request.user.is_WFM():
        try:
            user = CustomUser.objects.get(id=id)
            user.delete()
            success = True
        except CustomUser.DoesNotExist:
            logging.error(
                f"|Failed| Delete A User |id:{id}| Exception: CustomUser does not exist."
            )
        except Exception as e:
            logging.error(f"|Failed| Delete A User |id:{id}| Exception:{e}")
    else:
        errorMessage = "User does not have permission to delete"
    response = JsonResponse(
        {
            "success": success,
            "message": successMessage if success is True else errorMessage,
        }
    )
    return response


class viewUserJson(BaseDatatableView):
    model = CustomUser

    # Define the columns you want to display
    columns = [
        "name",
        "email",
        "is_active",
        "is_staff",
        "groups",
        "actions",
    ]

    # Define the order columns, make sure they match the columns order
    order_columns = [
        "name",
        "email",
        "is_active",
        "is_staff",
        "",
        "",  # Add an empty string for the actions column
    ]

    def get_initial_queryset(self):
        # Log the request
        if self.request.user.is_WFM():
            return CustomUser.objects.exclude(
                groups__name__in=["Employee", "Supervisor"]
            ).order_by("name")

    def filter_queryset(self, qs):
        # Handle POST parameters for filtering the queryset
        # Handle search parameter from Datatable
        searchValue = self.request.GET.get("search[value]", None)
        if searchValue:
            # Define the fields you want to search on
            searchFields = ["name", "email", "is_active", "is_staff", "groups__name"]

            # Create a Q object to dynamically construct the filter conditions
            searchFilter = Q()
            for field in searchFields:
                searchFilter |= Q(**{f"{field}__icontains": searchValue})

            # Apply the search filter to the queryset
            qs = qs.filter(searchFilter)
        return qs

    def prepare_results(self, qs):
        data = []
        for item in qs:
            groupNames = ", ".join([group.name for group in item.groups.all()])
            # Fetch the related field and use it directly in the data dictionary
            row = {
                "name": item.name.upper(),
                "email": item.email,
                "is_active": item.is_active,
                "is_staff": item.is_staff,
                "groups": groupNames,
                "actions": self.get_actions_html(item),
            }
            data.append(row)
        return data

    def render_column(self, row, column):
        return super(viewSkillJson, self).render_column(row, column)

    def get_actions_html(self, item):
        edit_url = reverse(PageInfoCollection.USER_EDIT.urlName, args=[item.id])
        reject_url = reverse(PageInfoCollection.USER_DELETE.urlName, args=[item.id])

        edit_link = format_html(
            '<a href="{}" class="btn btn-success">Edit</a>',
            edit_url,
        )
        reject_link = format_html(
            '<a href="{}" data-toggle="modal" data-target="#rejectModal" class="btn btn-danger">Delete</a>',
            reject_url,
        )

        return format_html("{} {}", edit_link, reject_link)
